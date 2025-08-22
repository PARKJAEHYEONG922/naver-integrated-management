"""
PowerLink 분석기 엑셀 내보내기 모듈
기존 통합관리프로그램과 동일한 형식으로 엑셀 파일 생성
"""
import openpyxl
from openpyxl.styles import Font, Alignment, NamedStyle
from typing import Dict
from datetime import datetime

from src.foundation.logging import get_logger
from src.desktop.common_log import log_manager
from .models import KeywordAnalysisResult, BidPosition

logger = get_logger("features.powerlink_analyzer.excel_export")


class PowerLinkExcelExporter:
    """PowerLink 분석 결과 엑셀 내보내기 클래스"""
    
    def __init__(self):
        pass
    
    def _setup_number_formats(self, workbook):
        """숫자 포맷 스타일 설정"""
        # 천 단위 콤마 스타일 (숫자로 저장하되 콤마 표시)
        if 'number_comma' not in workbook.named_styles:
            number_style = NamedStyle(name='number_comma')
            number_style.number_format = '#,##0'
            workbook.add_named_style(number_style)
        
        # 퍼센트 스타일
        if 'percent_style' not in workbook.named_styles:
            percent_style = NamedStyle(name='percent_style')
            percent_style.number_format = '0.00"%"'
            workbook.add_named_style(percent_style)
        
        return workbook
    
    def export_to_excel(self, keywords_data, file_path: str, session_name: str = ""):
        """
        PowerLink 분석 결과를 엑셀 파일로 내보내기
        
        Args:
            keywords_data: 키워드 분석 결과 딕셔너리 (KeywordAnalysisResult 또는 dict)
            file_path: 저장할 엑셀 파일 경로
            session_name: 세션명 (로그용)
        """
        try:
            if not keywords_data:
                raise Exception("저장할 분석 데이터가 없습니다.")
            
            # 데이터 정규화 (dict인 경우 KeywordAnalysisResult로 변환)
            normalized_data = {}
            for keyword, data in keywords_data.items():
                if isinstance(data, dict):
                    # dict인 경우 KeywordAnalysisResult 객체로 변환
                    normalized_data[keyword] = self._dict_to_result(data)
                else:
                    # 이미 KeywordAnalysisResult인 경우 그대로 사용
                    normalized_data[keyword] = data
            
            workbook = openpyxl.Workbook()
            
            # 숫자 포맷 스타일 설정
            workbook = self._setup_number_formats(workbook)
            
            # 모바일 분석결과 시트 생성
            self._create_mobile_sheet(workbook, normalized_data)
            
            # PC 분석결과 시트 생성
            self._create_pc_sheet(workbook, normalized_data)
            
            # 파일 저장
            workbook.save(file_path)
            
            log_message = f"PowerLink 엑셀 파일 생성 완료"
            if session_name:
                log_message += f" ({session_name})"
            log_message += f": {file_path}"
            
            log_manager.add_log(log_message, "success")
            logger.info(log_message)
            
        except Exception as e:
            error_message = f"PowerLink 엑셀 파일 생성 실패: {str(e)}"
            log_manager.add_log(error_message, "error")
            logger.error(error_message)
            raise
    
    def _create_mobile_sheet(self, workbook: openpyxl.Workbook, keywords_data: Dict[str, KeywordAnalysisResult]):
        """모바일 분석결과 시트 생성"""
        # 첫 번째 시트를 모바일로 설정
        mobile_sheet = workbook.active
        mobile_sheet.title = "모바일분석결과"
        
        # 모바일 헤더 설정 (5위까지 순위별 광고비 포함)
        mobile_headers = [
            "키워드", "월검색량", "모바일클릭수", "모바일클릭률", 
            "1p노출수", "1등광고비", "최소노출가격", "추천순위", ""  # 빈 칸
        ]
        
        # 1위부터 5위까지 광고비 헤더 추가
        for i in range(1, 6):
            mobile_headers.append(f"{i}위광고비")
        
        # 헤더 스타일 적용
        for col, header in enumerate(mobile_headers, 1):
            cell = mobile_sheet.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
        
        # 모바일 데이터 추가 (추천순위 순으로 정렬)
        sorted_mobile = sorted(
            keywords_data.values(), 
            key=lambda x: x.mobile_recommendation_rank if x.mobile_recommendation_rank > 0 else 999
        )
        
        for row_idx, result in enumerate(sorted_mobile, 2):
            # 기본 데이터
            mobile_sheet.cell(row=row_idx, column=1, value=result.keyword)
            
            # 월검색량 (Mobile 검색량, 숫자로 저장, 콤마 표시)
            monthly_volume_cell = mobile_sheet.cell(row=row_idx, column=2, value=result.mobile_search_volume)
            monthly_volume_cell.style = 'number_comma'
            
            mobile_sheet.cell(row=row_idx, column=3, value=result.mobile_clicks)
            
            # 클릭률 (숫자로 저장, 퍼센트 표시)
            ctr_cell = mobile_sheet.cell(row=row_idx, column=4, value=result.mobile_ctr)
            ctr_cell.style = 'percent_style'
            
            mobile_sheet.cell(row=row_idx, column=5, value=result.mobile_first_page_positions)
            
            # 1등광고비 (숫자로 저장, 콤마 표시)
            first_bid_cell = mobile_sheet.cell(row=row_idx, column=6, value=result.mobile_first_position_bid)
            first_bid_cell.style = 'number_comma'
            
            # 최소노출가격 (숫자로 저장, 콤마 표시)
            min_bid_cell = mobile_sheet.cell(row=row_idx, column=7, value=result.mobile_min_exposure_bid)
            min_bid_cell.style = 'number_comma'
            
            # 추천순위 (숫자로 저장)
            if result.mobile_recommendation_rank > 0:
                mobile_sheet.cell(row=row_idx, column=8, value=result.mobile_recommendation_rank)
            else:
                mobile_sheet.cell(row=row_idx, column=8, value="-")
            
            # 9번 컬럼은 빈 칸 (건너뜀)
            
            # 10번 컬럼부터 순위별 입찰가 (5위까지만) - 숫자로 저장, 콤마 표시
            if hasattr(result, 'mobile_bid_positions') and result.mobile_bid_positions:
                for idx, bid_pos in enumerate(result.mobile_bid_positions[:5]):
                    bid_cell = mobile_sheet.cell(row=row_idx, column=10+idx, value=bid_pos.bid_price)
                    bid_cell.style = 'number_comma'
        
        # 컬럼 너비 자동 조정
        self._auto_adjust_columns(mobile_sheet)
    
    def _create_pc_sheet(self, workbook: openpyxl.Workbook, keywords_data: Dict[str, KeywordAnalysisResult]):
        """PC 분석결과 시트 생성"""
        # PC 시트 생성
        pc_sheet = workbook.create_sheet("PC분석결과")
        
        # PC 헤더 설정 (10위까지 순위별 광고비 포함)
        pc_headers = [
            "키워드", "월검색량", "PC클릭수", "PC클릭률", 
            "1p노출수", "1등광고비", "최소노출가격", "추천순위", ""  # 빈 칸
        ]
        
        # 1위부터 10위까지 광고비 헤더 추가
        for i in range(1, 11):
            pc_headers.append(f"{i}위광고비")
        
        # 헤더 스타일 적용
        for col, header in enumerate(pc_headers, 1):
            cell = pc_sheet.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
        
        # PC 데이터 추가 (추천순위 순으로 정렬)
        sorted_pc = sorted(
            keywords_data.values(), 
            key=lambda x: x.pc_recommendation_rank if x.pc_recommendation_rank > 0 else 999
        )
        
        for row_idx, result in enumerate(sorted_pc, 2):
            # 기본 데이터
            pc_sheet.cell(row=row_idx, column=1, value=result.keyword)
            
            # 월검색량 (PC 검색량, 숫자로 저장, 콤마 표시)
            monthly_volume_cell = pc_sheet.cell(row=row_idx, column=2, value=result.pc_search_volume)
            monthly_volume_cell.style = 'number_comma'
            
            pc_sheet.cell(row=row_idx, column=3, value=result.pc_clicks)
            
            # 클릭률 (숫자로 저장, 퍼센트 표시)
            ctr_cell = pc_sheet.cell(row=row_idx, column=4, value=result.pc_ctr)
            ctr_cell.style = 'percent_style'
            
            pc_sheet.cell(row=row_idx, column=5, value=result.pc_first_page_positions)
            
            # 1등광고비 (숫자로 저장, 콤마 표시)
            first_bid_cell = pc_sheet.cell(row=row_idx, column=6, value=result.pc_first_position_bid)
            first_bid_cell.style = 'number_comma'
            
            # 최소노출가격 (숫자로 저장, 콤마 표시)
            min_bid_cell = pc_sheet.cell(row=row_idx, column=7, value=result.pc_min_exposure_bid)
            min_bid_cell.style = 'number_comma'
            
            # 추천순위 (숫자로 저장)
            if result.pc_recommendation_rank > 0:
                pc_sheet.cell(row=row_idx, column=8, value=result.pc_recommendation_rank)
            else:
                pc_sheet.cell(row=row_idx, column=8, value="-")
            
            # 9번 컬럼은 빈 칸 (건너뜀)
            
            # 10번 컬럼부터 순위별 입찰가 (10위까지만) - 숫자로 저장, 콤마 표시
            if hasattr(result, 'pc_bid_positions') and result.pc_bid_positions:
                for idx, bid_pos in enumerate(result.pc_bid_positions[:10]):
                    bid_cell = pc_sheet.cell(row=row_idx, column=10+idx, value=bid_pos.bid_price)
                    bid_cell.style = 'number_comma'
        
        # 컬럼 너비 자동 조정
        self._auto_adjust_columns(pc_sheet)
    
    def _auto_adjust_columns(self, sheet):
        """컬럼 너비 자동 조정"""
        try:
            for column in sheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                # 최소 10, 최대 30 범위로 제한
                adjusted_width = min(max(max_length + 2, 10), 30)
                sheet.column_dimensions[column_letter].width = adjusted_width
        except Exception as e:
            logger.warning(f"컬럼 너비 자동 조정 실패: {e}")
    
    def _dict_to_result(self, data_dict: dict) -> KeywordAnalysisResult:
        """딕셔너리를 KeywordAnalysisResult 객체로 변환"""
        from datetime import datetime
        
        # BidPosition 객체들 복원
        def restore_bid_positions(bid_data):
            if not bid_data:
                return []
            
            positions = []
            for item in bid_data:
                if isinstance(item, dict):
                    positions.append(BidPosition(
                        position=item.get('position', 0),
                        bid_price=item.get('bid_price', 0)
                    ))
                else:
                    # 이미 BidPosition 객체인 경우
                    positions.append(item)
            return positions
        
        # analyzed_at 필드 처리
        analyzed_at = data_dict.get('analyzed_at')
        if analyzed_at and isinstance(analyzed_at, str):
            try:
                analyzed_at = datetime.fromisoformat(analyzed_at)
            except:
                analyzed_at = datetime.now()
        elif not analyzed_at:
            analyzed_at = datetime.now()
        
        return KeywordAnalysisResult(
            keyword=data_dict.get('keyword', ''),
            pc_search_volume=data_dict.get('pc_search_volume', 0),
            mobile_search_volume=data_dict.get('mobile_search_volume', 0),
            pc_clicks=data_dict.get('pc_clicks', 0.0),
            pc_ctr=data_dict.get('pc_ctr', 0.0),
            pc_first_page_positions=data_dict.get('pc_first_page_positions', 0),
            pc_first_position_bid=data_dict.get('pc_first_position_bid', 0),
            pc_min_exposure_bid=data_dict.get('pc_min_exposure_bid', 0),
            pc_bid_positions=restore_bid_positions(data_dict.get('pc_bid_positions', [])),
            mobile_clicks=data_dict.get('mobile_clicks', 0.0),
            mobile_ctr=data_dict.get('mobile_ctr', 0.0),
            mobile_first_page_positions=data_dict.get('mobile_first_page_positions', 0),
            mobile_first_position_bid=data_dict.get('mobile_first_position_bid', 0),
            mobile_min_exposure_bid=data_dict.get('mobile_min_exposure_bid', 0),
            mobile_bid_positions=restore_bid_positions(data_dict.get('mobile_bid_positions', [])),
            pc_recommendation_rank=data_dict.get('pc_recommendation_rank', 0),
            mobile_recommendation_rank=data_dict.get('mobile_recommendation_rank', 0),
            analyzed_at=analyzed_at
        )


# 전역 익스포터 인스턴스
powerlink_excel_exporter = PowerLinkExcelExporter()