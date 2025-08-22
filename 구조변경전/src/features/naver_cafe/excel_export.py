"""
네이버 카페 DB 추출기 엑셀 내보내기 모듈
원본과 동일한 엑셀 저장 기능 제공
"""
import csv
from datetime import datetime
from pathlib import Path
from typing import List, Dict, Any

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from PySide6.QtWidgets import QFileDialog, QMessageBox

from src.foundation.logging import get_logger
from src.desktop.common_log import log_manager
from .models import ExtractedUser
from .config import CAFE_EXTRACTION_CONFIG

logger = get_logger("features.naver_cafe.excel_export")


class CafeExcelExporter:
    """카페 추출 데이터 엑셀 내보내기 클래스"""
    
    def __init__(self):
        self.default_font = Font(name='맑은 고딕', size=10)
        self.header_font = Font(name='맑은 고딕', size=11, bold=True)
        self.header_fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
        self.center_alignment = Alignment(horizontal='center', vertical='center')
    
    def export_to_excel(self, users_data: List[List[str]], parent_widget=None) -> bool:
        """
        엑셀로 내보내기 - 원본과 동일한 방식 (자동 폴더 열기 포함)
        
        Args:
            users_data: 사용자 데이터 리스트 [번호, 사용자ID, 닉네임, 추출시간]
            parent_widget: 부모 위젯 (다이얼로그용)
            
        Returns:
            bool: 성공 여부
        """
        if not users_data:
            if parent_widget:
                QMessageBox.information(parent_widget, "정보", "내보낼 데이터가 없습니다.")
            return False
        
        # 파일 저장 대화상자
        file_path, _ = QFileDialog.getSaveFileName(
            parent_widget,
            "저장할 파일명을 입력하세요",
            f"네이버카페_추출결과_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
            "Excel files (*.xlsx)"
        )
        
        if not file_path:
            return False
        
        try:
            # 워크북 생성
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "추출된 사용자"
            
            # 헤더 작성
            headers = ["번호", "사용자 ID", "닉네임", "추출 시간"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = self.header_font
                cell.fill = self.header_fill
                cell.alignment = self.center_alignment
            
            # 데이터 작성
            for row_idx, row_data in enumerate(users_data, 2):
                for col_idx, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.font = self.default_font
                    if col_idx == 1:  # 번호 컬럼 가운데 정렬
                        cell.alignment = self.center_alignment
            
            # 컬럼 너비 자동 조정
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # 파일 저장
            wb.save(file_path)
            
            filename = Path(file_path).name
            
            # 저장 완료 다이얼로그와 자동 폴더 열기
            self._show_save_completion_dialog(
                parent_widget,
                "저장 완료",
                f"카페 추출 데이터가 성공적으로 Excel 파일로 저장되었습니다.\n\n파일명: {filename}\n사용자 수: {len(users_data)}명",
                file_path
            )
            
            log_manager.add_log(f"Excel 파일 저장 완료: {filename}", "success")
            return True
            
        except ImportError:
            if parent_widget:
                QMessageBox.warning(parent_widget, "오류", "openpyxl 라이브러리가 필요합니다.\npip install openpyxl")
            logger.error("openpyxl 라이브러리 없음")
            return False
        except Exception as e:
            if parent_widget:
                QMessageBox.critical(parent_widget, "오류", f"엑셀 저장 중 오류가 발생했습니다.\n{str(e)}")
            logger.error(f"엑셀 저장 오류: {e}")
            return False
    
    def export_to_meta_csv(self, users_data: List[List[str]], parent_widget=None) -> bool:
        """
        Meta CSV로 내보내기 - 원본과 동일한 방식 (3개 도메인으로 저장)
        
        Args:
            users_data: 사용자 데이터 리스트 [번호, 사용자ID, 닉네임, 추출시간]
            parent_widget: 부모 위젯 (다이얼로그용)
            
        Returns:
            bool: 성공 여부
        """
        if not users_data:
            if parent_widget:
                QMessageBox.information(parent_widget, "정보", "내보낼 데이터가 없습니다.")
            return False
        
        # 파일 저장 대화상자
        file_path, _ = QFileDialog.getSaveFileName(
            parent_widget,
            "Meta CSV 파일 저장",
            f"네이버카페_Meta_{datetime.now().strftime('%Y%m%d_%H%M')}.csv",
            "CSV files (*.csv)"
        )
        
        if not file_path:
            return False
        
        try:
            # 사용자 ID만 추출 (인덱스 1)
            user_ids = [row_data[1] for row_data in users_data if len(row_data) >= 2]
            
            # 이메일 형태로 변환 (원본과 동일하게 3개 도메인)
            email1 = [user_id + '@naver.com' for user_id in user_ids]
            email2 = [user_id + '@gmail.com' for user_id in user_ids]
            email3 = [user_id + '@daum.net' for user_id in user_ids]
            
            # pandas DataFrame 생성 (원본과 동일한 방식)
            import pandas as pd
            df_meta = pd.DataFrame({
                'email': email1,
                'email2': email2,
                'email3': email3
            })
            
            # 헤더를 email,email,email 로 통일 (원본과 동일)
            df_meta.columns = ['email', 'email', 'email']
            
            # CSV 저장
            df_meta.to_csv(file_path, index=False)
            
            filename = Path(file_path).name
            
            # 저장 완료 다이얼로그와 자동 폴더 열기
            self._show_save_completion_dialog(
                parent_widget,
                "Meta CSV 저장 완료",
                f"Meta 광고용 CSV 파일이 성공적으로 저장되었습니다.\n\n파일명: {filename}\n사용자 ID: {len(user_ids)}개\n생성된 이메일: {len(user_ids)*3}개\n(@naver.com, @gmail.com, @daum.net)",
                file_path
            )
            
            log_manager.add_log(f"Meta CSV 파일 저장 완료: {filename} (사용자 {len(user_ids)}명 → 이메일 {len(user_ids)*3}개)", "success")
            return True
            
        except Exception as e:
            if parent_widget:
                QMessageBox.critical(parent_widget, "오류", f"CSV 저장 중 오류가 발생했습니다.\n{str(e)}")
            logger.error(f"CSV 저장 오류: {e}")
            return False
    
    def _show_save_completion_dialog(self, parent_widget, title: str, message: str, file_path: str):
        """저장 완료 다이얼로그 표시 및 자동 폴더 열기 (모던한 다이얼로그 사용)"""
        from src.toolbox.ui_kit.modern_dialog import ModernSaveCompletionDialog
        
        # 모던한 저장 완료 다이얼로그 표시
        ModernSaveCompletionDialog.show_save_completion(
            parent_widget, 
            title, 
            message, 
            file_path
        )
    


# 전역 인스턴스
cafe_excel_exporter = CafeExcelExporter()