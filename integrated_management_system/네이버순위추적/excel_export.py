"""
엑셀 저장 기능
기존 통합관리프로그램의 엑셀 저장 기능을 새 구조에 맞게 구현
"""
from datetime import datetime
from typing import List, Optional
import os

from src.foundation.logging import get_logger
from src.foundation.db import get_db
from .service import rank_tracking_service

logger = get_logger("features.rank_tracking.excel_export")


class RankTrackingExcelExporter:
    """순위 추적 엑셀 내보내기 클래스"""
    
    def __init__(self):
        pass
    
    def export_project_to_excel(self, project_id: int, file_path: str) -> bool:
        """단일 프로젝트를 엑셀로 저장 (간단한 형태)"""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            
            # 프로젝트 정보 조회
            project = rank_tracking_service.get_project_by_id(project_id)
            if not project:
                logger.error(f"프로젝트를 찾을 수 없습니다: {project_id}")
                return False
            
            # 키워드 정보 조회
            keywords = rank_tracking_service.get_project_keywords(project_id)
            
            # 워크북 생성
            workbook = openpyxl.Workbook()
            worksheet = workbook.active
            worksheet.title = "프로젝트 정보"
            
            # 프로젝트 정보 작성
            project_info = [
                ['프로젝트 ID', str(project.id)],
                ['상품 ID', project.product_id],
                ['상품명', project.current_name],
                ['카테고리', project.category or '-'],
                ['가격', f"{project.price:,}원" if project.price else '-'],
                ['스토어명', project.store_name or '-'],
                ['URL', project.product_url],
                ['생성일시', str(project.created_at) if project.created_at else '-']
            ]
            
            # 프로젝트 정보 섹션
            worksheet.append(['📊 프로젝트 정보'])
            worksheet.append([])
            for info in project_info:
                worksheet.append(info)
            
            worksheet.append([])
            worksheet.append(['🔤 키워드 목록'])
            worksheet.append(['키워드', '카테고리', '월검색량', '활성 상태'])
            
            # 키워드 데이터 추가
            for keyword in keywords:
                # UI와 동일한 월검색량 표시 규칙 적용
                monthly_vol = keyword.monthly_volume
                if monthly_vol is None or monthly_vol == -1:
                    volume_display = "-"  # 검색량 못가져왔을 때
                elif monthly_vol == 0:
                    volume_display = 0  # 검색량이 0일 때
                else:
                    volume_display = monthly_vol  # 정상 검색량
                
                worksheet.append([
                    keyword.keyword,
                    keyword.category or '-',
                    volume_display,
                    '활성' if keyword.is_active else '비활성'
                ])
            
            # 파일 저장
            workbook.save(file_path)
            workbook.close()
            
            logger.info(f"엑셀 파일 생성 완료: {file_path}")
            return True
            
        except Exception as e:
            logger.error(f"엑셀 저장 중 오류: {e}")
            return False
    
    def export_multiple_projects_to_excel(self, project_ids: List[int], file_path: str) -> bool:
        """여러 프로젝트를 엑셀로 저장 (단일 저장을 시트별로 분할)"""
        try:
            import openpyxl
            
            # 워크북 생성
            workbook = openpyxl.Workbook()
            
            # 기본 시트 제거
            if 'Sheet' in workbook.sheetnames:
                del workbook['Sheet']
            
            # 각 프로젝트별로 시트 생성
            for i, project_id in enumerate(project_ids, 1):
                try:
                    logger.info(f"다중 프로젝트 내보내기: 프로젝트 {project_id} 처리 시작")
                    
                    # 임시 파일에 단일 프로젝트 저장 (temp 폴더 사용)
                    import tempfile
                    import os
                    temp_dir = tempfile.gettempdir()
                    temp_file = os.path.join(temp_dir, f"temp_project_{project_id}.xlsx")
                    success = self.export_ranking_history_to_excel(project_id, temp_file)
                    
                    if success:
                        logger.info(f"프로젝트 {project_id} 임시 파일 생성 성공: {temp_file}")
                        
                        # 임시 파일을 워크북에 시트로 추가
                        temp_workbook = openpyxl.load_workbook(temp_file)
                        source_sheet = temp_workbook.active
                        
                        # 임시 파일의 데이터 확인
                        row_count = source_sheet.max_row
                        col_count = source_sheet.max_column
                        logger.info(f"프로젝트 {project_id} 임시 파일 데이터: {row_count}행 x {col_count}열")
                        
                        # 헤더와 첫 번째 데이터 행 확인
                        if row_count >= 12:  # 헤더가 12번째 행
                            header_row = [cell.value for cell in source_sheet[12]]
                            logger.info(f"프로젝트 {project_id} 헤더: {header_row[:5]}...")  # 처음 5개만
                            
                            if row_count >= 13:  # 첫 번째 데이터 행
                                data_row = [cell.value for cell in source_sheet[13]]
                                logger.info(f"프로젝트 {project_id} 첫 데이터: {data_row[:5]}...")  # 처음 5개만
                        
                        # 새 시트 생성
                        target_sheet = workbook.create_sheet(title=f"Sheet{i}")
                        
                        # 데이터와 스타일 안전한 복사
                        for row in source_sheet.iter_rows():
                            for cell in row:
                                new_cell = target_sheet.cell(row=cell.row, column=cell.column, value=cell.value)
                                
                                # 스타일 안전한 복사 (StyleProxy 오류 방지)
                                try:
                                    if cell.font:
                                        from openpyxl.styles import Font
                                        new_cell.font = Font(
                                            name=cell.font.name,
                                            size=cell.font.size,
                                            bold=cell.font.bold,
                                            italic=cell.font.italic,
                                            color=cell.font.color
                                        )
                                except:
                                    pass
                                    
                                try:
                                    if cell.fill:
                                        from openpyxl.styles import PatternFill
                                        new_cell.fill = PatternFill(
                                            start_color=cell.fill.start_color,
                                            end_color=cell.fill.end_color,
                                            fill_type=cell.fill.fill_type
                                        )
                                except:
                                    pass
                                    
                                try:
                                    if cell.alignment:
                                        from openpyxl.styles import Alignment
                                        new_cell.alignment = Alignment(
                                            horizontal=cell.alignment.horizontal,
                                            vertical=cell.alignment.vertical
                                        )
                                except:
                                    pass
                                    
                                try:
                                    if cell.number_format:
                                        new_cell.number_format = cell.number_format
                                except:
                                    pass
                        
                        # 컬럼 너비 복사
                        for col_letter, dimension in source_sheet.column_dimensions.items():
                            target_sheet.column_dimensions[col_letter].width = dimension.width
                        
                        # 임시 워크북 완전히 닫기
                        temp_workbook.close()
                        
                        # 파일 잠금 해제를 위한 대기
                        import time
                        time.sleep(0.3)
                        
                        # 임시 파일 안전하게 삭제 (조용히)
                        import os
                        try:
                            if os.path.exists(temp_file):
                                os.remove(temp_file)
                                logger.debug(f"임시 파일 삭제 성공: {temp_file}")
                        except:
                            # 삭제 실패는 로그 출력하지 않음 (시스템이 알아서 정리함)
                            pass
                        
                        logger.info(f"프로젝트 {project_id} 시트 생성 완료")
                    else:
                        logger.error(f"프로젝트 {project_id} 단일 저장 실패")
                        
                except Exception as e:
                    logger.error(f"프로젝트 {project_id} 처리 중 오류: {e}")
                    continue
            
            if len(workbook.sheetnames) == 0:
                logger.error("저장할 프로젝트 데이터가 없습니다")
                workbook.close()
                return False
            
            # 파일 저장
            workbook.save(file_path)
            workbook.close()
            
            logger.info(f"다중 프로젝트 엑셀 파일 생성 완료: {file_path}")
            return True
            
        except Exception as e:
            logger.error(f"다중 프로젝트 엑셀 저장 중 오류: {e}")
            return False
    
    def export_ranking_history_to_excel(self, project_id: int, file_path: str, ranking_data: list = None) -> bool:
        """순위 이력을 엑셀로 저장 (사진과 똑같은 형태)"""
        try:
            # 프로젝트 정보 조회
            project = rank_tracking_service.get_project_by_id(project_id)
            if not project:
                logger.error(f"프로젝트를 찾을 수 없습니다: {project_id}")
                return False
            
            # 키워드 정보 조회
            keywords = rank_tracking_service.get_project_keywords(project_id)
            if not keywords:
                logger.error("키워드가 없습니다")
                return False
            
            # 프로젝트 순위 개요 데이터 사용
            overview = rank_tracking_service.get_project_overview(project_id)
            all_dates = overview.get('dates', [])[:10]  # 최대 10개 날짜
            keywords_data = overview.get('keywords', {})
            
            logger.info(f"디버깅: 키워드 수 = {len(keywords)}")
            logger.info(f"디버깅: 전체 날짜 수 = {len(all_dates)}, 날짜들 = {all_dates}")
            
            # 키워드별 순위 데이터 구성
            keyword_ranking_data = []
            for keyword_obj in keywords:
                keyword_id = keyword_obj.id
                keyword_data = keywords_data.get(keyword_id, {})
                rankings = keyword_data.get('rankings', {})
                
                logger.info(f"디버깅: 키워드 '{keyword_obj.keyword}' (ID: {keyword_id}) 순위 이력 수 = {len(rankings)}")
                
                # 날짜별 순위 매핑 (overview 데이터 형식에 맞춤)
                rank_by_date = {}
                for date in all_dates:
                    if date in rankings:
                        rank = rankings[date].get('rank', 999)
                        rank_by_date[date] = rank
                        logger.info(f"디버깅: 순위 데이터 - 키워드: {keyword_obj.keyword}, 날짜: {date}, 순위: {rank}")
                
                keyword_ranking_data.append({
                    'keyword': keyword_obj.keyword,
                    'category': keyword_obj.category or '-',
                    'monthly_volume': keyword_obj.monthly_volume if keyword_obj.monthly_volume is not None else -1,
                    'rank_by_date': rank_by_date
                })
            
            # 날짜 정렬 (최신순) 및 형식 변환
            sorted_dates = []
            formatted_dates = []
            for date in all_dates:
                # 날짜를 MM/DD HH:MM 형식으로 변환
                try:
                    if isinstance(date, str):
                        dt = datetime.fromisoformat(date.replace('Z', '+00:00'))
                        formatted_date = dt.strftime("%m/%d %H:%M")
                        sorted_dates.append(date)  # 원본 날짜 (키 매칭용)
                        formatted_dates.append(formatted_date)  # 표시용 날짜
                        logger.info(f"디버깅: 날짜 변환 - {date} -> {formatted_date}")
                except Exception as e:
                    logger.warning(f"디버깅: 날짜 변환 실패 - {date}: {e}")
                    continue
            
            logger.info(f"디버깅: 최종 날짜 수 = {len(sorted_dates)}, 변환된 날짜들 = {formatted_dates}")
            
            # 엑셀 데이터 구성
            excel_data = []
            
            # 1. 기본정보 섹션 (사진과 똑같이)
            excel_data.extend([
                [f"📊 {project.current_name}", "", "", "", "", "", "", "", ""],
                ["", "", "", "", "", "", "", "", ""],
                ["상품 ID", project.product_id, "", "", "", "", "", "", ""],
                ["상품명", project.current_name, "", "", "", "", "", "", ""],
                ["스토어명", project.store_name or "-", "", "", "", "", "", "", ""],
                ["가격", f"{project.price:,}원" if project.price else "-", "", "", "", "", "", "", ""],
                ["카테고리", project.category or "-", "", "", "", "", "", "", ""],
                ["등록일", self._format_date(project.created_at) if project.created_at else "-", "", "", "", "", "", "", ""],
                ["", "", "", "", "", "", "", "", ""],
                ["", "", "", "", "", "", "", "", ""],
                ["🔍 키워드 순위 현황", "", "", "", "", "", "", "", ""]
            ])
            
            # 2. 키워드 순위 테이블 헤더 (표시용 날짜 사용)
            header_row = ["키워드", "카테고리", "월검색량"]
            header_row.extend(formatted_dates)  # 변환된 날짜들 추가
            excel_data.append(header_row)
            
            # 3. 키워드별 순위 데이터
            for kw_data in keyword_ranking_data:
                # UI와 동일한 월검색량 표시 규칙 적용
                monthly_vol = kw_data['monthly_volume']
                if monthly_vol == -1:
                    volume_display = "-"  # 검색량 못가져왔을 때
                elif monthly_vol == 0:
                    volume_display = "0"  # 검색량이 0일 때
                else:
                    volume_display = f"{monthly_vol:,}"  # 정상 검색량
                
                data_row = [
                    kw_data['keyword'],
                    kw_data['category'],
                    volume_display
                ]
                
                # 각 날짜별 순위 추가 (원본 날짜로 키 매칭)
                for date in sorted_dates:
                    rank = kw_data['rank_by_date'].get(date, "")
                    if rank:
                        if rank == 999 or rank > 200:  # RANK_OUT_OF_RANGE 또는 200위 초과
                            data_row.append("200+")
                        else:
                            data_row.append(f"{rank}위")
                    else:
                        data_row.append("")
                
                excel_data.append(data_row)
            
            # 엑셀 파일 생성
            success = self._create_excel_file(file_path, excel_data)
            
            if success:
                logger.info(f"순위 이력 엑셀 파일 생성 완료: {file_path}")
            else:
                logger.error("순위 이력 엑셀 파일 생성 실패")
            
            return success
            
        except Exception as e:
            logger.error(f"순위 이력 엑셀 저장 중 오류: {e}")
            return False


    def get_default_filename(self, project_id: Optional[int] = None) -> str:
        """기본 파일명 생성"""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        
        if project_id:
            project = rank_tracking_service.get_project_by_id(project_id)
            if project:
                # 파일명에 사용할 수 없는 문자 제거
                safe_name = project.current_name.replace('/', '_').replace('\\', '_')
                safe_name = safe_name.replace(':', '_').replace('*', '_')
                safe_name = safe_name.replace('?', '_').replace('"', '_')
                safe_name = safe_name.replace('<', '_').replace('>', '_')
                safe_name = safe_name.replace('|', '_')
                
                return f"순위이력_{safe_name}_{timestamp}.xlsx"
        
        return f"순위이력_데이터_{timestamp}.xlsx"
    
    def _format_date(self, date_value):
        """날짜 형식을 안전하게 변환"""
        try:
            if isinstance(date_value, str):
                # 문자열인 경우 datetime으로 변환
                dt = datetime.fromisoformat(date_value.replace('Z', '+00:00'))
                return dt.strftime("%Y-%m-%d")
            elif hasattr(date_value, 'strftime'):
                # datetime 객체인 경우
                return date_value.strftime("%Y-%m-%d")
            else:
                # 기타 경우는 문자열로 변환
                return str(date_value)
        except Exception as e:
            logger.warning(f"날짜 형식 변환 실패: {e}")
            return str(date_value) if date_value else "-"
    
    def _create_excel_file(self, file_path: str, excel_data: list) -> bool:
        """엑셀 파일 생성"""
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            
            # 워크북 생성
            workbook = openpyxl.Workbook()
            worksheet = workbook.active
            worksheet.title = "순위이력"
            
            # 데이터 입력 및 정렬 가능하도록 처리
            for row_idx, row_data in enumerate(excel_data, 1):
                for col_idx, cell_value in enumerate(row_data, 1):
                    # 월검색량과 순위 컬럼은 숫자로 저장하여 정렬 가능하게 함
                    if row_idx > 12 and col_idx == 3:  # 월검색량 컬럼
                        # 월검색량을 UI와 동일하게 처리
                        try:
                            if isinstance(cell_value, str):
                                if cell_value == "-":
                                    # 검색량 못가져왔을 때는 "-" 문자열로 저장
                                    cell = worksheet.cell(row=row_idx, column=col_idx, value="-")
                                elif cell_value == "0":
                                    # 검색량 0일 때는 숫자 0으로 저장
                                    cell = worksheet.cell(row=row_idx, column=col_idx, value=0)
                                    cell.number_format = '#,##0'
                                elif cell_value.replace(',', '').isdigit():
                                    numeric_value = int(cell_value.replace(',', ''))
                                    cell = worksheet.cell(row=row_idx, column=col_idx, value=numeric_value)
                                    cell.number_format = '#,##0'
                                else:
                                    cell = worksheet.cell(row=row_idx, column=col_idx, value=cell_value)
                            elif isinstance(cell_value, int):
                                cell = worksheet.cell(row=row_idx, column=col_idx, value=cell_value)
                                cell.number_format = '#,##0'
                            else:
                                cell = worksheet.cell(row=row_idx, column=col_idx, value=cell_value)
                        except:
                            cell = worksheet.cell(row=row_idx, column=col_idx, value=cell_value)
                    elif row_idx > 12 and col_idx > 3:  # 순위 컬럼들
                        # 순위를 숫자로 변환하여 정렬 가능하게 함
                        try:
                            if isinstance(cell_value, str):
                                if "200+" in cell_value:
                                    cell = worksheet.cell(row=row_idx, column=col_idx, value=201)  # 정렬용
                                    # 200+ 표시를 위한 number format 설정
                                    cell.number_format = '"200+"'
                                elif "위" in cell_value:
                                    rank_num = int(cell_value.replace("위", ""))
                                    cell = worksheet.cell(row=row_idx, column=col_idx, value=rank_num)
                                    # N위 표시를 위한 number format 설정
                                    cell.number_format = '0"위"'
                                else:
                                    cell = worksheet.cell(row=row_idx, column=col_idx, value=cell_value)
                            else:
                                cell = worksheet.cell(row=row_idx, column=col_idx, value=cell_value)
                        except:
                            cell = worksheet.cell(row=row_idx, column=col_idx, value=cell_value)
                    else:
                        cell = worksheet.cell(row=row_idx, column=col_idx, value=cell_value)
                    
                    # 스타일 적용
                    if row_idx == 1:  # 제목 행
                        cell.font = Font(size=14, bold=True)
                        cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
                        cell.font = Font(color="FFFFFF", size=14, bold=True)
                    elif row_idx == 11:  # 키워드 순위 현황 헤더 (11번째 행)
                        cell.font = Font(size=12, bold=True)
                        cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
                    elif row_idx == 12:  # 테이블 헤더 (12번째 행이 실제 헤더)
                        cell.font = Font(color="FFFFFF", bold=True)
                        cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                    elif row_idx > 12:  # 데이터 행 (12번째 행 이후)
                        if col_idx <= 3:  # 키워드, 카테고리, 월검색량 컬럼
                            cell.alignment = Alignment(horizontal="left", vertical="center")
                        else:  # 순위 컬럼들
                            cell.alignment = Alignment(horizontal="center", vertical="center")
                            # 순위에 따른 색상 적용 (숫자 값 기준)
                            if isinstance(cell.value, (int, float)):
                                rank_num = int(cell.value)
                                if rank_num <= 10:
                                    cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # 연한 초록색
                                elif rank_num <= 50:
                                    cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")  # 연한 노란색
                                elif rank_num <= 200:
                                    cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # 연한 빨간색
                                else:  # 200+
                                    cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # 연한 빨간색
            
            # 컬럼 너비 설정
            for col_idx in range(1, worksheet.max_column + 1):
                column_letter = worksheet.cell(row=1, column=col_idx).column_letter
                
                if col_idx == 1:  # 키워드 컬럼
                    worksheet.column_dimensions[column_letter].width = 20
                elif col_idx == 2:  # 카테고리 컬럼
                    worksheet.column_dimensions[column_letter].width = 30
                elif col_idx == 3:  # 월검색량 컬럼
                    worksheet.column_dimensions[column_letter].width = 12
                else:  # 순위 컬럼들 (기존 대비 1.5배)
                    worksheet.column_dimensions[column_letter].width = 15  # 기본 10 → 15로 1.5배
            
            # 파일 저장
            workbook.save(file_path)
            return True
            
        except Exception as e:
            logger.error(f"엑셀 파일 생성 중 오류: {e}")
            return False


# 전역 인스턴스
rank_tracking_excel_exporter = RankTrackingExcelExporter()