"""
엑셀 파일 생성/읽기/내보내기
Excel 관련 모든 기능을 담당
"""
import os
import subprocess
from pathlib import Path
from typing import List, Dict, Any, Optional
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

from src.foundation.exceptions import FileError
from src.foundation.logging import get_logger


logger = get_logger("toolbox.excel")


class ExcelExporter:
    """Excel 파일 내보내기 클래스"""
    
    def __init__(self):
        self.default_font = Font(name='맑은 고딕', size=10)
        self.header_font = Font(name='맑은 고딕', size=11, bold=True)
        self.header_fill = PatternFill(start_color='E6E6FA', end_color='E6E6FA', fill_type='solid')
        self.center_alignment = Alignment(horizontal='center', vertical='center')
    
    def export_keywords(self, 
                       data: List[Dict[str, Any]], 
                       file_path: str,
                       sheet_name: str = "키워드 분석") -> bool:
        """
        키워드 데이터를 Excel로 내보내기
        
        Args:
            data: 키워드 데이터 리스트
            file_path: 저장할 파일 경로
            sheet_name: 시트 이름
        
        Returns:
            bool: 성공 여부
        """
        try:
            if not data:
                logger.warning("내보낼 데이터가 없습니다")
                return False
            
            # DataFrame 생성
            df = pd.DataFrame(data)
            
            # 컬럼명 한글화
            column_mapping = {
                'keyword': '키워드',
                'category': '카테고리',
                'search_volume': '월간 검색량',
                'total_products': '상품 수',
                'competition_strength': '경쟁 강도'
            }
            
            df = df.rename(columns=column_mapping)
            
            # Excel 파일 생성
            wb = Workbook()
            ws = wb.active
            ws.title = sheet_name
            
            # 데이터 추가
            for r in dataframe_to_rows(df, index=False, header=True):
                ws.append(r)
            
            # 스타일 적용
            self._apply_styles(ws, len(df.columns))
            
            # 파일 저장
            wb.save(file_path)
            logger.info(f"Excel 파일 저장 완료: {file_path}")
            return True
            
        except Exception as e:
            logger.error(f"Excel 내보내기 실패: {e}")
            raise FileError(f"Excel 파일 생성 실패: {e}")
    
    def export_ranking_results(self,
                              data: List[Dict[str, Any]],
                              file_path: str,
                              sheet_name: str = "순위 추적") -> bool:
        """
        순위 추적 결과를 Excel로 내보내기
        
        Args:
            data: 순위 데이터 리스트
            file_path: 저장할 파일 경로
            sheet_name: 시트 이름
        
        Returns:
            bool: 성공 여부
        """
        try:
            if not data:
                logger.warning("내보낼 데이터가 없습니다")
                return False
            
            # DataFrame 생성
            df = pd.DataFrame(data)
            
            # 컬럼명 한글화
            column_mapping = {
                'keyword': '키워드',
                'rank': '순위',
                'previous_rank': '이전 순위',
                'rank_change': '순위 변동',
                'check_date': '확인일시'
            }
            
            # 존재하는 컬럼만 매핑
            existing_columns = {k: v for k, v in column_mapping.items() if k in df.columns}
            df = df.rename(columns=existing_columns)
            
            # Excel 파일 생성
            wb = Workbook()
            ws = wb.active
            ws.title = sheet_name
            
            # 데이터 추가
            for r in dataframe_to_rows(df, index=False, header=True):
                ws.append(r)
            
            # 스타일 적용
            self._apply_styles(ws, len(df.columns))
            
            # 파일 저장
            wb.save(file_path)
            logger.info(f"Excel 파일 저장 완료: {file_path}")
            return True
            
        except Exception as e:
            logger.error(f"Excel 내보내기 실패: {e}")
            raise FileError(f"Excel 파일 생성 실패: {e}")
    
    def _apply_styles(self, worksheet, column_count: int):
        """워크시트에 스타일 적용"""
        try:
            # 헤더 스타일
            for col in range(1, column_count + 1):
                cell = worksheet.cell(row=1, column=col)
                cell.font = self.header_font
                cell.fill = self.header_fill
                cell.alignment = self.center_alignment
            
            # 데이터 스타일
            for row in worksheet.iter_rows(min_row=2):
                for cell in row:
                    cell.font = self.default_font
                    if cell.column == 1:  # 첫 번째 컬럼 (키워드) 좌측 정렬
                        cell.alignment = Alignment(horizontal='left', vertical='center')
                    elif cell.column == 2:  # 두 번째 컬럼 (카테고리) - 줄바꿈 허용
                        cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
                    elif cell.column == 3:  # 세 번째 컬럼 (월간 검색량) - 천단위 콤마
                        cell.alignment = Alignment(horizontal='right', vertical='center')
                        if cell.value and isinstance(cell.value, (int, float)):
                            cell.number_format = '#,##0'
                    elif cell.column == 4:  # 네 번째 컬럼 (상품 수) - 천단위 콤마
                        cell.alignment = Alignment(horizontal='right', vertical='center')
                        if cell.value and isinstance(cell.value, (int, float)):
                            cell.number_format = '#,##0'
                    elif cell.column == 5:  # 다섯 번째 컬럼 (경쟁 강도) - 소수점 2자리
                        cell.alignment = Alignment(horizontal='right', vertical='center')
                        if cell.value and isinstance(cell.value, (int, float)):
                            cell.number_format = '0.00'
                    else:  # 나머지 컬럼 중앙 정렬
                        cell.alignment = self.center_alignment
            
            # 컬럼 너비 설정 (기존 통합관리프로그램과 동일)
            column_widths = {
                1: 20,   # 키워드
                2: 50,   # 카테고리 (줄바꿈을 위해 넓게)
                3: 15,   # 월간 검색량
                4: 15,   # 상품 수
                5: 12    # 경쟁 강도
            }
            
            for col_num, width in column_widths.items():
                if col_num <= column_count:
                    column_letter = worksheet.cell(row=1, column=col_num).column_letter
                    worksheet.column_dimensions[column_letter].width = width
            
            # 행 높이 설정 (카테고리 줄바꿈을 위해)
            for row_num in range(2, worksheet.max_row + 1):
                category_cell = worksheet.cell(row=row_num, column=2)
                if category_cell.value and '\n' in str(category_cell.value):
                    # 카테고리에 줄바꿈이 있으면 행 높이 늘리기
                    worksheet.row_dimensions[row_num].height = 30
                
        except Exception as e:
            logger.warning(f"스타일 적용 중 오류: {e}")

    def create_workbook_with_sheets(self, file_path: str, sheets_data: Dict[str, Dict[str, Any]]) -> bool:
        """
        여러 시트가 포함된 워크북 생성
        
        Args:
            file_path: 저장할 파일 경로
            sheets_data: {sheet_name: {'headers': [...], 'data': [...]}} 형태의 데이터
            
        Returns:
            bool: 성공 여부
        """
        try:
            # 워크북 생성
            wb = Workbook()
            
            # 기본 시트 제거
            if 'Sheet' in wb.sheetnames:
                wb.remove(wb['Sheet'])
            
            # 각 시트 생성
            for sheet_name, sheet_info in sheets_data.items():
                ws = wb.create_sheet(title=sheet_name)
                
                headers = sheet_info.get('headers', [])
                data = sheet_info.get('data', [])
                
                if headers:
                    # 헤더 추가
                    for col, header in enumerate(headers, 1):
                        cell = ws.cell(row=1, column=col, value=header)
                        cell.font = self.header_font
                        cell.fill = self.header_fill
                        cell.alignment = self.center_alignment
                
                # 데이터 추가
                for row, row_data in enumerate(data, 2):  # 헤더 다음 행부터 시작
                    for col, value in enumerate(row_data, 1):
                        cell = ws.cell(row=row, column=col, value=value)
                        cell.font = self.default_font
                        if col == 1:  # 첫 번째 컬럼은 왼쪽 정렬
                            cell.alignment = Alignment(horizontal='left', vertical='center')
                        else:
                            cell.alignment = self.center_alignment
                
                # 컬럼 너비 자동 조정
                for col in ws.columns:
                    max_length = 0
                    column = col[0].column_letter
                    for cell in col:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)  # 최대 50자로 제한
                    ws.column_dimensions[column].width = adjusted_width
            
            # 파일 저장
            wb.save(file_path)
            logger.info(f"워크북 생성 완료: {file_path}")
            return True
            
        except Exception as e:
            logger.error(f"워크북 생성 실패: {e}")
            return False


class ExcelReader:
    """Excel 파일 읽기 클래스"""
    
    def read_keywords(self, file_path: str, sheet_name: Optional[str] = None) -> List[Dict[str, Any]]:
        """
        Excel 파일에서 키워드 데이터 읽기
        
        Args:
            file_path: Excel 파일 경로
            sheet_name: 시트 이름 (None이면 첫 번째 시트)
        
        Returns:
            List[Dict]: 키워드 데이터 리스트
        """
        try:
            # Excel 파일 읽기
            df = pd.read_excel(file_path, sheet_name=sheet_name)
            
            # 데이터 변환
            data = df.to_dict('records')
            
            logger.info(f"Excel 파일 읽기 완료: {len(data)}개 레코드")
            return data
            
        except Exception as e:
            logger.error(f"Excel 읽기 실패: {e}")
            raise FileError(f"Excel 파일 읽기 실패: {e}")


def open_file_location(file_path: str) -> bool:
    """
    파일 위치 열기
    
    Args:
        file_path: 파일 경로
    
    Returns:
        bool: 성공 여부
    """
    try:
        if not os.path.exists(file_path):
            logger.error(f"파일이 존재하지 않습니다: {file_path}")
            return False
        
        # Windows에서 파일 탐색기로 파일 위치 열기
        if os.name == 'nt':  # Windows
            subprocess.run(['explorer', '/select,', file_path], check=True)
        else:
            # macOS/Linux의 경우
            import platform
            if platform.system() == 'Darwin':  # macOS
                subprocess.run(['open', '-R', file_path], check=True)
            else:  # Linux
                subprocess.run(['xdg-open', os.path.dirname(file_path)], check=True)
        
        logger.info(f"파일 위치 열기 성공: {file_path}")
        return True
        
    except Exception as e:
        logger.error(f"파일 위치 열기 실패: {e}")
        return False


# 전역 인스턴스
excel_exporter = ExcelExporter()
excel_reader = ExcelReader()