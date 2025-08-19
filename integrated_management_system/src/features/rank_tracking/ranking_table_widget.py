"""
순위 테이블 위젯 - 키워드 순위 관리 및 표시
기존 UI와 완전 동일한 스타일 및 기능
"""
from datetime import datetime
from typing import Optional, Dict
from PySide6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QPushButton, QLabel,
    QTreeWidget, QTreeWidgetItem, QFrame, QCheckBox,
    QDialog, QGridLayout, QApplication
)
from PySide6.QtCore import Qt, Signal, QTimer
from PySide6.QtGui import QFont, QColor

from src.toolbox.ui_kit import ModernStyle, SortableTreeWidgetItem
from src.desktop.common_log import log_manager
from src.toolbox.ui_kit import ModernTextInputDialog, ModernSaveCompletionDialog
from src.foundation.logging import get_logger

from .worker import RankingCheckWorker, ranking_worker_manager, keyword_info_worker_manager
from .excel_export import rank_tracking_excel_exporter
from .adapters import format_date, format_date_with_time, format_rank_display, get_rank_color, format_monthly_volume, get_category_match_color
from .view_model import ranking_table_view_model
# repository.py 제거됨 - Foundation DB 사용으로 단순화

logger = get_logger("features.rank_tracking.ranking_table_widget")





class RankingTableWidget(QWidget):
    """순위 테이블 위젯 - 기존과 완전 동일"""
    
    project_updated = Signal()  # 프로젝트 업데이트 시그널
    
    def __init__(self):
        super().__init__()
        self.current_project_id = None
        self.current_project = None
        self.selected_projects = []  # 다중 선택된 프로젝트들
        self.setup_ui()
        
        # 워커 매니저 시그널 연결
        ranking_worker_manager.progress_updated.connect(self.on_progress_updated)
        ranking_worker_manager.keyword_rank_updated.connect(self.on_keyword_rank_updated)
        ranking_worker_manager.ranking_finished.connect(self.on_ranking_finished)
        
        # 키워드 정보 워커 매니저 시그널 연결
        keyword_info_worker_manager.progress_updated.connect(self.on_keyword_info_progress_updated)
        keyword_info_worker_manager.category_updated.connect(self.on_keyword_category_updated)
        keyword_info_worker_manager.volume_updated.connect(self.on_keyword_volume_updated)
        keyword_info_worker_manager.keyword_info_finished.connect(self.on_keyword_info_finished)
    
    def setup_ui(self):
        """UI 구성 - 원본과 완전 동일"""
        layout = QVBoxLayout()
        layout.setContentsMargins(10, 10, 10, 10)
        layout.setSpacing(10)
        
        # 상품 정보 영역 추가
        self.info_widget = self.create_product_info_widget()
        layout.addWidget(self.info_widget)
        
        # 테이블 상단 버튼들
        button_layout = QHBoxLayout()
        
        # 키워드 삭제 버튼
        self.delete_keywords_button = QPushButton("🗑️ 선택 키워드 삭제")
        self.delete_keywords_button.clicked.connect(self.delete_selected_keywords)
        self.delete_keywords_button.setStyleSheet("""
            QPushButton {
                background-color: #dc3545;
                color: white;
                border: none;
                padding: 8px 12px;
                border-radius: 6px;
                font-weight: 600;
            }
            QPushButton:hover {
                background-color: #c82333;
            }
            QPushButton:disabled {
                background-color: #D1D5DB;
                color: white;
            }
        """)
        self.delete_keywords_button.setEnabled(False)
        button_layout.addWidget(self.delete_keywords_button)
        
        # 진행상황 표시를 버튼 옆에 배치 (원본과 동일)
        self.progress_frame = QFrame()
        self.progress_frame.setVisible(False)
        progress_layout = QHBoxLayout()  # 가로 배치로 변경
        progress_layout.setContentsMargins(5, 5, 5, 5)  # 여백 최소화
        progress_layout.setSpacing(8)  # 간격을 8px로 줄임
        
        from PySide6.QtWidgets import QProgressBar, QSizePolicy
        
        self.progress_label = QLabel("작업 진행 중...")
        self.progress_label.setFont(QFont("맑은 고딕", 10))  # 폰트 크기 줄임
        self.progress_label.setStyleSheet("color: #007ACC; font-weight: 500;")
        self.progress_label.setSizePolicy(QSizePolicy.Fixed, QSizePolicy.Fixed)  # 크기 고정
        progress_layout.addWidget(self.progress_label)
        
        self.progress_bar = QProgressBar()
        self.progress_bar.setRange(0, 100)
        self.progress_bar.setFixedHeight(16)  # 높이 제한
        self.progress_bar.setFixedWidth(150)  # 폭 제한
        self.progress_bar.setVisible(False)  # 단계 진행시에만 표시
        self.progress_bar.setSizePolicy(QSizePolicy.Fixed, QSizePolicy.Fixed)  # 크기 고정
        progress_layout.addWidget(self.progress_bar)
        
        progress_layout.addStretch()  # 오른쪽에 늘어나는 공간 추가
        
        self.progress_frame.setLayout(progress_layout)
        button_layout.addWidget(self.progress_frame)
        
        button_layout.addStretch()
        layout.addLayout(button_layout)
        
        # 순위 테이블
        self.ranking_table = QTreeWidget()
        self.setup_ranking_table()
        layout.addWidget(self.ranking_table)
        
        # 하단 버튼들
        self.setup_buttons(layout)
        
        self.setLayout(layout)
        
        # 강제 새로고침 메서드 추가
        self.force_refresh_ranking_table = self._force_refresh_ranking_table
        self.rebuild_ranking_table = self._rebuild_ranking_table
    
    def create_product_info_widget(self) -> QWidget:
        """상품 정보 영역 생성"""
        from PySide6.QtWidgets import QFrame, QGridLayout
        from PySide6.QtGui import QFont
        
        widget = QFrame()
        widget.setStyleSheet(f"""
            QFrame {{
                background-color: #F8F9FA;
                border-radius: 8px;
                border: 1px solid {ModernStyle.COLORS['border']};
                padding: 15px;
            }}
        """)
        
        layout = QGridLayout()
        layout.setSpacing(12)
        layout.setVerticalSpacing(10)
        
        # 헤더 스타일 설정 (원본과 동일)
        header_style = """
            QLabel {
                color: #495057;
                font-weight: 600;
                padding: 2px 0px;
                border: none;
            }
        """
        
        value_style = """
            QLabel { 
                padding: 4px 8px; 
                border: 1px solid transparent;
                border-radius: 4px;
                background-color: #FFFFFF;
                color: #212529;
            } 
            QLabel:hover { 
                border: 1px solid #DEE2E6; 
                background-color: #F8F9FA;
            }
        """
        
        # Row 0: 상품ID (새로고침 버튼 포함)
        product_id_header = QLabel("상품ID")
        product_id_header.setFont(QFont("맑은 고딕", 11))
        product_id_header.setStyleSheet(header_style)
        layout.addWidget(product_id_header, 0, 0)
        
        # 상품ID 행에 수평 레이아웃 생성 (ID + 새로고침 버튼)
        from PySide6.QtWidgets import QHBoxLayout
        product_id_layout = QHBoxLayout()
        
        self.product_id_label = QLabel("-")
        self.product_id_label.setFont(QFont("맑은 고딕", 11))
        self.product_id_label.setStyleSheet(value_style)
        product_id_layout.addWidget(self.product_id_label)
        
        # 새로고침 버튼 (심플 디자인)
        from PySide6.QtWidgets import QPushButton
        self.refresh_product_button = QPushButton("⟲")
        self.refresh_product_button.setToolTip("상품 정보 새로고침")
        self.refresh_product_button.setFixedSize(28, 28)
        self.refresh_product_button.setStyleSheet("""
            QPushButton {
                background-color: #F3F4F6;
                color: #6B7280;
                border: 1px solid #D1D5DB;
                border-radius: 6px;
                font-size: 14px;
                font-weight: 500;
                padding: 0px;
            }
            QPushButton:hover {
                background-color: #E5E7EB;
                color: #374151;
                border-color: #9CA3AF;
            }
            QPushButton:pressed {
                background-color: #D1D5DB;
                color: #111827;
                border-color: #6B7280;
            }
            QPushButton:disabled {
                background-color: #F9FAFB;
                color: #D1D5DB;
                border-color: #F3F4F6;
            }
        """)
        self.refresh_product_button.clicked.connect(self.refresh_product_info)
        product_id_layout.addWidget(self.refresh_product_button)
        product_id_layout.addStretch()  # 오른쪽 여백
        
        # 레이아웃을 위젯으로 만들어서 그리드에 추가
        product_id_widget = QWidget()
        product_id_widget.setLayout(product_id_layout)
        layout.addWidget(product_id_widget, 0, 1)
        
        # Row 1: 상품명
        product_name_header = QLabel("상품명")
        product_name_header.setFont(QFont("맑은 고딕", 11))
        product_name_header.setStyleSheet(header_style)
        layout.addWidget(product_name_header, 1, 0)
        
        # 상품명 행에 수평 레이아웃 생성 (상품명 + 변경사항 버튼)
        product_name_layout = QHBoxLayout()
        
        self.product_name_label = QLabel("-")
        self.product_name_label.setFont(QFont("맑은 고딕", 11))
        self.product_name_label.setStyleSheet(value_style)
        product_name_layout.addWidget(self.product_name_label)
        
        # 변경사항 버튼 (기존 통합관리프로그램과 동일한 디자인)
        self.changes_button = QPushButton("📝")
        self.changes_button.setToolTip("프로젝트 변경사항 보기")
        self.changes_button.setFixedSize(28, 28)
        self.changes_button.setStyleSheet("""
            QPushButton {
                background-color: #F3F4F6;
                color: #6B7280;
                border: 1px solid #D1D5DB;
                border-radius: 6px;
                font-size: 14px;
                font-weight: 500;
                padding: 0px;
            }
            QPushButton:hover {
                background-color: #E5E7EB;
                color: #374151;
                border-color: #9CA3AF;
            }
            QPushButton:pressed {
                background-color: #D1D5DB;
                color: #111827;
                border-color: #6B7280;
            }
            QPushButton:disabled {
                background-color: #F9FAFB;
                color: #D1D5DB;
                border-color: #F3F4F6;
            }
        """)
        self.changes_button.clicked.connect(self.show_project_changes)
        product_name_layout.addWidget(self.changes_button)
        product_name_layout.addStretch()  # 오른쪽 여백
        
        # 레이아웃을 위젯으로 만들어서 그리드에 추가
        product_name_widget = QWidget()
        product_name_widget.setLayout(product_name_layout)
        layout.addWidget(product_name_widget, 1, 1)
        
        # Row 2: 스토어명
        store_name_header = QLabel("스토어명")
        store_name_header.setFont(QFont("맑은 고딕", 11))
        store_name_header.setStyleSheet(header_style)
        layout.addWidget(store_name_header, 2, 0)
        
        self.store_name_label = QLabel("-")
        self.store_name_label.setFont(QFont("맑은 고딕", 11))
        self.store_name_label.setStyleSheet(value_style)
        layout.addWidget(self.store_name_label, 2, 1)
        
        # Row 3: 가격
        price_header = QLabel("가격")
        price_header.setFont(QFont("맑은 고딕", 11))
        price_header.setStyleSheet(header_style)
        layout.addWidget(price_header, 3, 0)
        
        self.price_label = QLabel("-")
        self.price_label.setFont(QFont("맑은 고딕", 11))
        self.price_label.setStyleSheet(value_style)
        layout.addWidget(self.price_label, 3, 1)
        
        # Row 4: 카테고리
        category_header = QLabel("카테고리")
        category_header.setFont(QFont("맑은 고딕", 11))
        category_header.setStyleSheet(header_style)
        layout.addWidget(category_header, 4, 0)
        
        self.category_label = QLabel("-")
        self.category_label.setFont(QFont("맑은 고딕", 11))
        self.category_label.setStyleSheet(value_style)
        layout.addWidget(self.category_label, 4, 1)
        
        # 컬럼 너비 설정 (헤더는 고정폭, 값은 유동적)
        layout.setColumnStretch(0, 0)  # 헤더 컬럼은 고정
        layout.setColumnStretch(1, 1)  # 값 컬럼은 늘어남
        
        # 마지막 확인 시간 (카드 외부에 별도 표시)
        self.last_check_label = QLabel("마지막 확인: -")
        self.last_check_label.setFont(QFont("맑은 고딕", 12, QFont.Bold))
        self.last_check_label.setStyleSheet("color: #495057; margin-top: 8px; font-weight: bold;")
        layout.addWidget(self.last_check_label, 5, 0, 1, 2)  # Row 5로 이동 - 두 컬럼에 걸쳐서 표시
        
        widget.setLayout(layout)
        return widget
    
    def refresh_product_info(self):
        """상품 정보 새로고침"""
        if not self.current_project_id:
            return
        
        # 버튼 로딩 상태로 변경
        self.refresh_product_button.setEnabled(False)
        self.refresh_product_button.setText("⟳")
        
        # 1초 후 버튼 상태 복원 (임시 구현)
        from PySide6.QtCore import QTimer
        QTimer.singleShot(1000, lambda: (
            self.refresh_product_button.setEnabled(True),
            self.refresh_product_button.setText("⟲")
        ))
    
    def setup_ranking_table(self):
        """순위 테이블 설정 (원본과 완전 동일)"""
        # 기본 헤더 설정
        self.ranking_table.setHeaderLabels([
            "", "키워드", "카테고리", "월검색량"
        ])
        
        # 헤더 보이기
        self.ranking_table.header().setVisible(True)
        
        # 정렬 기능 활성화 (Qt 기본 정렬 사용)
        self.ranking_table.setSortingEnabled(True)
        self.ranking_table.header().setSectionsClickable(True)
        
        # 정렬 후 헤더 체크박스 위치 재조정을 위한 시그널 연결
        self.ranking_table.header().sortIndicatorChanged.connect(self.on_sort_changed)
        
        # 헤더 우클릭 메뉴 설정
        self.ranking_table.header().setContextMenuPolicy(Qt.CustomContextMenu)
        self.ranking_table.header().customContextMenuRequested.connect(self.show_header_context_menu)
        
        # 헤더 설정 (원본과 동일)
        header = self.ranking_table.header()
        header.setDefaultSectionSize(100)  # 기본 컬럼 너비
        header.setMinimumSectionSize(50)   # 최소 컬럼 너비
        
        # 컬럼 너비 설정 (원본과 완전 동일)
        self.ranking_table.setColumnWidth(0, 80)       # 체크박스 (원본과 동일)
        self.ranking_table.setColumnWidth(1, 200)      # 키워드
        self.ranking_table.setColumnWidth(2, 180)      # 카테고리
        self.ranking_table.setColumnWidth(3, 100)      # 월검색량
        
        # 테이블 스타일 설정 (원본과 완전 동일)
        self.ranking_table.setStyleSheet(f"""
            QTreeWidget {{
                border: 1px solid {ModernStyle.COLORS['border']};
                border-radius: 8px;
                background-color: {ModernStyle.COLORS['bg_card']};
                alternate-background-color: {ModernStyle.COLORS['bg_primary']};
                font-size: 13px;
                gridline-color: {ModernStyle.COLORS['border']};
            }}
            QTreeWidget::item {{
                padding: 8px 4px;
                border-bottom: 1px solid {ModernStyle.COLORS['border']};
            }}
            QTreeWidget::item:selected {{
                background-color: {ModernStyle.COLORS['primary']};
                color: white;
            }}
            QHeaderView::section {{
                background-color: {ModernStyle.COLORS['bg_input']};
                border: none;
                border-right: 1px solid {ModernStyle.COLORS['border']};
                border-bottom: 2px solid {ModernStyle.COLORS['border']};
                padding: 8px;
                font-weight: 600;
                color: {ModernStyle.COLORS['text_primary']};
            }}
        """)
    
    def on_sort_changed(self):
        """정렬 변경 시 헤더 체크박스 위치 재조정"""
        if hasattr(self, 'header_checkbox'):
            self.position_header_checkbox()
    
    def show_header_context_menu(self, position):
        """헤더 우클릭 컨텍스트 메뉴 표시"""
        if not self.current_project:
            return
            
        header = self.ranking_table.header()
        column = header.logicalIndexAt(position)
        
        # 날짜 컬럼인지 확인 (컬럼 4번 이후가 날짜 컬럼)
        if column < 4:  # 체크박스, 키워드, 카테고리, 월검색량 컬럼은 제외
            return
            
        # 헤더 텍스트에서 날짜 추출
        header_item = self.ranking_table.headerItem()
        if header_item and column < header_item.columnCount():
            column_text = header_item.text(column)
            if not column_text or column_text == "-":
                return
                
            # 컨텍스트 메뉴 생성
            from PySide6.QtWidgets import QMenu
            
            context_menu = QMenu(self)
            delete_action = context_menu.addAction(f"🗑️ {column_text} 날짜 데이터 삭제")
            delete_action.triggered.connect(lambda: self.delete_date_column_data(column, column_text))
            
            # 메뉴 표시
            global_pos = header.mapToGlobal(position)
            context_menu.exec(global_pos)
    
    def delete_date_column_data(self, column_index: int, date_text: str):
        """날짜 컬럼 데이터 삭제"""
        if not self.current_project:
            return
            
        from src.toolbox.ui_kit import ModernConfirmDialog
        
        # 확인 다이얼로그
        reply = ModernConfirmDialog.warning(
            self,
            "날짜 데이터 삭제",
            f"{date_text} 날짜의 모든 순위 데이터를 삭제하시겠습니까?\n\n• 해당 날짜 컬럼이 테이블에서 제거됩니다\n• 이 작업은 되돌릴 수 없습니다",
            "삭제", "취소"
        )
        
        if reply:
            try:
                # ViewModel을 통한 프로젝트 개요 조회
                overview = ranking_table_view_model.get_project_overview(self.current_project_id)
                dates = overview.get('dates', []) if overview else []
                
                # 헤더 인덱스에 맞는 날짜 찾기 (컬럼 4번부터 날짜)
                date_index = column_index - 4  # 컬럼 0,1,2,3은 체크박스, 키워드, 카테고리, 월검색량
                if 0 <= date_index < len(dates):
                    actual_date = dates[date_index]
                    logger.info(f"삭제할 실제 날짜: '{actual_date}'")
                    
                    # ViewModel을 통한 데이터베이스 삭제
                    success = ranking_table_view_model.delete_ranking_data_by_date(self.current_project_id, actual_date)
                    
                    if success:
                        log_manager.add_log(f"✅ {date_text} 날짜의 순위 데이터가 삭제되었습니다.", "success")
                        
                        # 해당 컬럼을 테이블에서 제거
                        # 모든 행의 해당 컬럼 데이터 제거
                        for i in range(self.ranking_table.topLevelItemCount()):
                            item = self.ranking_table.topLevelItem(i)
                            if item:
                                # 해당 컬럼의 데이터를 지우고 왼쪽으로 당기기
                                for col in range(column_index, self.ranking_table.columnCount() - 1):
                                    next_text = item.text(col + 1)
                                    next_data = item.data(col + 1, Qt.UserRole)
                                    next_color = item.foreground(col + 1)
                                    
                                    item.setText(col, next_text)
                                    item.setData(col, Qt.UserRole, next_data)
                                    item.setForeground(col, next_color)
                        
                        # 헤더도 왼쪽으로 당기기
                        for col in range(column_index, self.ranking_table.columnCount() - 1):
                            next_header = self.ranking_table.headerItem().text(col + 1)
                            self.ranking_table.headerItem().setText(col, next_header)
                        
                        # 마지막 컬럼 제거
                        self.ranking_table.setColumnCount(self.ranking_table.columnCount() - 1)
                        
                        # 테이블 업데이트
                        self.ranking_table.viewport().update()
                        self.ranking_table.repaint()
                    else:
                        log_manager.add_log(f"❌ {date_text} 날짜의 순위 데이터 삭제에 실패했습니다.", "error")
                        QMessageBox.information(self, "삭제 실패", "데이터베이스에서 해당 날짜의 데이터를 찾을 수 없거나 삭제에 실패했습니다.")
                else:
                    QMessageBox.information(self, "오류", "날짜 데이터를 찾을 수 없습니다.")
                    
            except Exception as e:
                log_manager.add_log(f"❌ 날짜 데이터 삭제 중 오류: {str(e)}", "error")
                QMessageBox.information(self, "오류", f"날짜 데이터 삭제 중 오류가 발생했습니다: {str(e)}")
    
    
    def apply_styles(self):
        """스타일 적용"""
        self.setStyleSheet(f"""
            QFrame {{
                border: 1px solid {ModernStyle.COLORS['border']};
                border-radius: 6px;
                padding: 10px;
                background-color: {ModernStyle.COLORS['bg_primary']};
            }}
            QTreeWidget {{
                border: 1px solid {ModernStyle.COLORS['border']};
                border-radius: 6px;
                background-color: {ModernStyle.COLORS['bg_primary']};
            }}
            QPushButton {{
                background-color: {ModernStyle.COLORS['primary']};
                color: white;
                border: none;
                padding: 10px 15px;
                border-radius: 6px;
                font-weight: 600;
                min-width: 100px;
            }}
            QPushButton:hover {{
                background-color: {ModernStyle.COLORS['primary_hover']};
            }}
            QPushButton:disabled {{
                background-color: {ModernStyle.COLORS['border']};
            }}
        """)
    
    def set_project(self, project):
        """프로젝트 설정"""
        logger.info(f"🔧 프로젝트 설정: ID={project.id}, 이름={getattr(project, 'current_name', 'N/A')}")
        logger.info(f"   - 프로젝트 카테고리: '{getattr(project, 'category', 'N/A')}'")
        
        # ViewModel에 현재 프로젝트 설정
        if ranking_table_view_model.set_current_project(project.id):
            self.current_project = project
            self.current_project_id = project.id
            self.update_project_info(project.id)
            logger.info(f"✅ 프로젝트 설정 완료: current_project_id={self.current_project_id}")
        else:
            logger.error(f"프로젝트 설정 실패: {project.id}")
        
        # 버튼 활성화 및 상태 업데이트
        if hasattr(self, 'add_keyword_button'):
            self.add_keyword_button.setEnabled(True)
        if hasattr(self, 'save_button'):
            self.save_button.setEnabled(True)
        
        # 순위 확인 버튼 상태는 해당 프로젝트의 실행 상태에 따라 결정
        self.update_button_state_from_project_status(project.id)
        
        # 진행률 표시 상태도 프로젝트에 따라 업데이트
        self.update_progress_display_from_project_status(project.id)
    
    def update_button_state_from_project_status(self, project_id):
        """프로젝트 상태에 따른 버튼 상태 업데이트"""
        if hasattr(self, 'check_button') and hasattr(self, 'stop_button'):
            is_running = ranking_worker_manager.is_ranking_in_progress(project_id)
            self.update_button_state_for_current_project(running=is_running)
            logger.info(f"프로젝트 {project_id} 버튼 상태 복원: 순위 확인 {'진행중' if is_running else '대기중'}")
    
    def update_progress_display_from_project_status(self, project_id):
        """프로젝트 상태에 따른 진행률 표시 업데이트"""
        logger.info(f"프로젝트 {project_id} 진행률 표시 업데이트 확인")
        
        current, total = ranking_worker_manager.get_current_progress(project_id)
        if current > 0 and total > 0:
            self.show_progress(f"순위 확인 중... ({current}/{total})", show_bar=True)
            percentage = int((current / total) * 100) if total > 0 else 0
            self.progress_bar.setValue(percentage)
            logger.info(f"✅ 프로젝트 {project_id} 진행률 복원: {current}/{total} ({percentage}%)")
        else:
            self.hide_progress()
            logger.info(f"프로젝트 {project_id} 진행률 없음 - 진행률바 숨김")
    
    def clear_project(self):
        """프로젝트 초기화 - 삭제 시 호출"""
        # 프로젝트 정보 초기화
        self.current_project = None
        self.current_project_id = None
        
        # 테이블 초기화
        if hasattr(self, 'ranking_table'):
            self.ranking_table.clear()
        
        # 모든 버튼 비활성화
        if hasattr(self, 'add_keyword_button'):
            self.add_keyword_button.setEnabled(False)
        if hasattr(self, 'check_button'):
            self.check_button.setEnabled(False)
        if hasattr(self, 'save_button'):
            self.save_button.setEnabled(False)
        if hasattr(self, 'delete_keywords_button'):
            self.delete_keywords_button.setEnabled(False)
        
        # 진행 상태 숨기기
        if hasattr(self, 'progress_frame'):
            self.progress_frame.setVisible(False)
    
    def update_project_info(self, project_id: int):
        """프로젝트 정보 업데이트"""
        self.current_project_id = project_id
        
        # ViewModel에서 현재 프로젝트 설정 및 정보 조회
        if ranking_table_view_model.set_current_project(project_id):
            project = ranking_table_view_model.get_current_project()
        else:
            project = None
        if not project:
            return
            
        # 상품 정보 표시
        self.product_id_label.setText(f"{project.product_id}")
        self.product_name_label.setText(f"{project.current_name}")
        
        # 스토어명 표시
        store_name = project.store_name if project.store_name else "정보 없음"
        self.store_name_label.setText(f"{store_name}")
        
        # 가격 표시 (천 단위 콤마 포함)
        if project.price and project.price > 0:
            self.price_label.setText(f"{project.price:,}원")
        else:
            self.price_label.setText("정보 없음")
        
        # 카테고리 표시
        category = getattr(project, 'category', '') or "-"
        self.category_label.setText(category)
        
        # 순위 현황 표시
        self.update_ranking_table(project_id)
    
    def update_ranking_table(self, project_id):
        """순위 테이블 업데이트 (진행 중인 순위 확인 상태 고려)"""
        try:
            # 현재 프로젝트에서 순위 확인이 진행 중인지 확인
            is_ranking_in_progress = ranking_worker_manager.is_ranking_in_progress(project_id)
            
            # 진행 중인 경우 진행률 상태만 복원하고 테이블은 정상 구성
            if is_ranking_in_progress:
                logger.info(f"프로젝트 {project_id}: 순위 확인 진행 중 - 진행 상태 복원하고 테이블 구성")
                self.update_progress_display_from_project_status(project_id)
            
            # ViewModel에서 테이블 데이터 준비
            table_data = ranking_table_view_model.prepare_table_data(project_id)
            if not table_data.get("success", False):
                logger.error(f"테이블 데이터 준비 실패: {table_data.get('message')}")
                return
            
            # ViewModel에서 준비된 데이터 사용
            headers = table_data["headers"]
            keywords_data = table_data["overview"].get("keywords", {}) if table_data["overview"] else {}
            dates = table_data["dates"]
            project_category_base = table_data["project_category_base"]
            
            # 키워드 데이터가 없으면 직접 키워드 목록에서 가져오기
            if not keywords_data:
                keywords = table_data["keywords"]
                for keyword in keywords:
                    keywords_data[keyword.id] = {
                        'keyword': keyword.keyword,
                        'category': keyword.category or '-',
                        'monthly_volume': keyword.monthly_volume if keyword.monthly_volume is not None else -1,
                        'search_volume': getattr(keyword, 'search_volume', None),
                        'is_active': True,
                        'rankings': {}
                    }
            
            # 날짜 정보는 이미 ViewModel에서 처리됨
            all_dates = dates
            
            # 현재 진행 중인 순위 확인이 있다면 해당 시간도 포함 (기존 로직 유지)
            current_time = ranking_worker_manager.get_current_time(project_id)
            if current_time and current_time not in all_dates:
                all_dates = [current_time] + all_dates
                # 헤더도 다시 업데이트 필요
                headers = table_data["headers"][:4]  # 기본 4개 헤더
                for date in all_dates:
                    headers.append(format_date(date))
                
            # 마지막 확인 시간 표시 (ViewModel에서 준비됨)
            self.last_check_label.setText(table_data["last_check_time"])
            
            # 테이블 완전 초기화 및 헤더 설정
            self.ranking_table.clear()
            self.ranking_table.setColumnCount(len(headers))
            self.ranking_table.setHeaderLabels(headers)
            
            # 헤더 체크박스 초기화
            if hasattr(self, 'header_checkbox'):
                try:
                    self.header_checkbox.setParent(None)
                    self.header_checkbox.deleteLater()
                    delattr(self, 'header_checkbox')
                except:
                    pass
            
            # 키워드가 없어도 헤더는 표시됨
            if not keywords_data:
                self.update_delete_button_state()
                return
            
            # 키워드별 행 추가
            for keyword_id, data in keywords_data.items():
                keyword = data['keyword']
                is_active = data.get('is_active', True)  # 기본값 True로 설정
                rankings = data.get('rankings', {})
                
                # 리스트로 아이템 데이터 준비
                row_data = ["", keyword]  # 첫 번째는 체크박스용 빈 문자열
                
                # 카테고리 추가
                category = data.get('category', '') or '-'
                row_data.append(category)
                
                # 월검색량
                search_vol = data.get('search_volume')
                monthly_vol = data.get('monthly_volume', -1)
                volume = search_vol or monthly_vol
                
                # 월검색량 포맷팅
                if volume == -1:
                    volume_text = "-"  # 아직 API 호출 안됨 (UI에서는 "-"으로 표시)
                else:
                    volume_text = format_monthly_volume(volume)
                row_data.append(volume_text)
                
                # 날짜별 순위 추가 (진행 중인 날짜 포함)
                for date in all_dates:
                    # 진행 중인 날짜인 경우 임시 저장된 순위 데이터 확인
                    current_time = ranking_worker_manager.get_current_time(project_id)
                    if date == current_time:
                        current_rankings = ranking_worker_manager.get_current_rankings(project_id)
                        if keyword_id in current_rankings:
                            rank = current_rankings[keyword_id]
                            rank_display = format_rank_display(rank)
                            row_data.append(rank_display)
                        else:
                            row_data.append("-")
                    else:
                        # 저장된 순위 데이터 확인
                        rank_data = rankings.get(date)
                        if rank_data and rank_data.get('rank') is not None:
                            rank_display = format_rank_display(rank_data['rank'])
                            row_data.append(rank_display)
                        else:
                            row_data.append("-")
                
                # 아이템 생성 및 추가 (SortableTreeWidgetItem 사용)
                item = SortableTreeWidgetItem(row_data)
                
                # 월검색량을 정렬용 데이터로 저장
                if volume_text == '-':
                    volume_for_sort = -1  # "-"는 가장 아래로 정렬
                else:
                    volume_for_sort = volume if volume is not None else 0
                item.setData(3, Qt.UserRole, volume_for_sort)
                
                # 날짜별 순위도 정렬용 숫자 데이터로 저장 + 색깔 적용
                for i, date in enumerate(all_dates):
                    column_index = 4 + i  # 4번째 컬럼부터 순위 데이터
                    
                    # 진행 중인 날짜인 경우 임시 저장된 순위 데이터 확인
                    current_time = ranking_worker_manager.get_current_time(project_id)
                    if date == current_time:
                        current_rankings = ranking_worker_manager.get_current_rankings(project_id)
                        if keyword_id in current_rankings:
                            actual_rank = current_rankings[keyword_id]
                            sort_rank = 201 if (actual_rank == 0 or actual_rank > 200) else actual_rank
                            item.setData(column_index, Qt.UserRole, sort_rank)
                        else:
                            item.setData(column_index, Qt.UserRole, 999)  # "-"는 가장 뒤로 정렬
                        continue
                        
                    rank_data = rankings.get(date)
                    if rank_data and rank_data.get('rank') is not None:
                        actual_rank = rank_data['rank']
                        # 0이나 200초과는 201로 저장 (정렬 시 맨 아래로)
                        sort_rank = 201 if (actual_rank == 0 or actual_rank > 200) else actual_rank
                        item.setData(column_index, Qt.UserRole, sort_rank)
                        
                        # 순위에 따른 색깔 설정
                        color = get_rank_color(actual_rank, "foreground")
                        item.setForeground(column_index, QColor(color))
                    else:
                        # 데이터가 없는 경우 202로 저장 (맨 아래로)
                        item.setData(column_index, Qt.UserRole, 202)
                        color = get_rank_color(0, "foreground")  # 데이터 없음
                        item.setForeground(column_index, QColor(color))
                
                self.ranking_table.addTopLevelItem(item)
                
                # 카테고리 색상 적용 (프로젝트 카테고리와 비교)
                if project_category_base and category != '-':
                    # 키워드 카테고리에서 괄호 앞 부분만 추출
                    keyword_category_clean = category.split('(')[0].strip()
                    color = get_category_match_color(project_category_base, keyword_category_clean)
                    item.setForeground(2, QColor(color))
                
                # 삭제 선택용 체크박스 추가
                checkbox = self._create_item_checkbox()
                self.ranking_table.setItemWidget(item, 0, checkbox)
                
                # 키워드 ID를 키워드 컬럼에 저장 (삭제 시 사용)
                item.setData(1, Qt.UserRole, keyword_id)
            
            # 월검색량 기준 내림차순 자동 정렬 (키워드가 있을 때만)
            if keywords_data:
                self.ranking_table.sortByColumn(3, Qt.DescendingOrder)
            
            # 헤더에 체크박스 추가
            self.setup_header_checkbox()
            
            # 삭제 버튼 상태 업데이트
            self.update_delete_button_state()
                
        except Exception as e:
            logger.error(f"순위 테이블 업데이트 실패: {e}")
    
    
    def _create_item_checkbox(self) -> QCheckBox:
        """아이템용 체크박스 생성 (공통 스타일 적용)"""
        checkbox = QCheckBox()
        checkbox.setChecked(False)
        checkbox.setStyleSheet("""
            QCheckBox {
                spacing: 0px;
                margin: 2px;
                padding: 1px;
            }
            QCheckBox::indicator {
                width: 14px;
                height: 14px;
                border: 2px solid #ccc;
                border-radius: 3px;
                background-color: white;
            }
            QCheckBox::indicator:checked {
                background-color: #dc3545;
                border-color: #dc3545;
            }
        """)
        checkbox.stateChanged.connect(lambda: self.update_delete_button_state())
        return checkbox
    
    def _create_header_checkbox(self) -> QCheckBox:
        """헤더용 체크박스 생성 (공통 스타일 적용)"""
        checkbox = QCheckBox()
        checkbox.setStyleSheet("""
            QCheckBox {
                spacing: 0px;
                margin: 2px;
                padding: 2px;
                border: none;
                background-color: transparent;
            }
            QCheckBox::indicator {
                width: 14px;
                height: 14px;
                border: 2px solid #ccc;
                border-radius: 3px;
                background-color: white;
                margin: 1px;
            }
            QCheckBox::indicator:checked {
                background-color: #dc3545;
                border-color: #dc3545;
            }
            QCheckBox::indicator:hover {
                border-color: #999999;
                background-color: #f8f9fa;
            }
        """)
        checkbox.stateChanged.connect(self.on_header_checkbox_changed)
        return checkbox
    
    def _get_primary_button_style(self, min_width: str = "100px") -> str:
        """기본 버튼 스타일 (파란색)"""
        return f"""
            QPushButton {{
                background-color: #007bff;
                color: white;
                border: none;
                padding: 8px 12px;
                border-radius: 6px;
                font-weight: 600;
                min-width: {min_width};
            }}
            QPushButton:hover {{
                background-color: #0056b3;
            }}
            QPushButton:disabled {{
                background-color: #D1D5DB;
                color: white;
            }}
        """
    
    def _get_danger_button_style(self, min_width: str = "80px") -> str:
        """위험 버튼 스타일 (빨간색)"""
        return f"""
            QPushButton {{
                background-color: #dc3545;
                color: white;
                border: none;
                padding: 8px 12px;
                border-radius: 6px;
                font-weight: 600;
                min-width: {min_width};
            }}
            QPushButton:hover {{
                background-color: #c82333;
            }}
            QPushButton:disabled {{
                background-color: #D1D5DB;
                color: white;
            }}
        """
    
    def _get_success_button_style(self, min_width: str = "100px") -> str:
        """성공 버튼 스타일 (녹색)"""
        return f"""
            QPushButton {{
                background-color: #10b981;
                color: white;
                border: none;
                padding: 8px 12px;
                border-radius: 6px;
                font-weight: 600;
                min-width: {min_width};
            }}
            QPushButton:hover {{
                background-color: #059669;
            }}
            QPushButton:disabled {{
                background-color: #D1D5DB;
                color: white;
            }}
        """
    
    def setup_header_checkbox(self):
        """QTreeWidget 헤더에 체크박스 추가 (원본과 완전 동일한 오버레이 방식)"""
        try:
            # 헤더용 체크박스 생성
            self.header_checkbox = self._create_header_checkbox()
            
            # QTreeWidget의 헤더뷰를 가져와서 첫 번째 섹션에 위젯 설정
            header = self.ranking_table.header()
            
            # 첫 번째 컬럼의 크기를 고정하고 위젯을 배치할 공간 확보
            from PySide6.QtWidgets import QHeaderView
            header.setSectionResizeMode(0, QHeaderView.Fixed)
            header.resizeSection(0, 80)  # 체크박스 컬럼 너비
            
            # 첫 번째 컬럼 헤더를 빈 문자열로 설정
            header_item = self.ranking_table.headerItem()
            if header_item:
                header_item.setText(0, "")
            
            # 헤더 위치에 overlay 방식으로 체크박스 배치
            self.position_header_checkbox()
            
        except Exception as e:
            logger.error(f"헤더 체크박스 설정 실패: {e}")
    
    def reset_header_checkbox(self):
        """헤더 체크박스를 완전히 재설정 (키워드 추가 후 호출)"""
        try:
            # 기존 헤더 체크박스 제거
            if hasattr(self, 'header_checkbox') and self.header_checkbox:
                try:
                    self.header_checkbox.deleteLater()
                    delattr(self, 'header_checkbox')
                except:
                    pass
            
            # 짧은 지연 후 새로 생성 (UI 업데이트 보장)
            QTimer.singleShot(100, self.setup_header_checkbox)
            
        except Exception as e:
            logger.error(f"헤더 체크박스 재설정 실패: {e}")

    def position_header_checkbox(self):
        """헤더 위치에 체크박스 오버레이 (원본과 완전 동일)"""
        try:
            if not hasattr(self, 'header_checkbox') or not self.header_checkbox:
                return
                
            # QTreeWidget의 헤더 영역 위치 계산
            header = self.ranking_table.header()
            
            # 안전한 위치 계산
            if header.sectionSize(0) <= 0:
                return
                
            header_rect = header.sectionViewportPosition(0), 0, header.sectionSize(0), header.height()
            
            # 체크박스를 헤더 위에 오버레이로 배치 (부모는 한번만 설정)
            if self.header_checkbox.parent() != header:
                self.header_checkbox.setParent(header)
            
            header_height = header.height()
            
            self.header_checkbox.setGeometry(
                header_rect[0] + 22,                                      # x: 왼쪽에서 22px 여백
                6,                                                        # y: 위쪽에서 6px 여백
                header_rect[2] - 20,                                      # width: 컬럼 너비에서 좌우 10px씩 여백
                header_height - 12                                        # height: 헤더 높이에서 위아래 6px씩 여백
            )
            self.header_checkbox.show()
            
        except Exception as e:
            logger.error(f"헤더 체크박스 위치 설정 실패: {e}")
    
    def on_header_checkbox_changed(self):
        """헤더 체크박스 상태 변경 (원본과 완전 동일)"""
        try:
            if not hasattr(self, 'header_checkbox'):
                return
                
            is_checked = self.header_checkbox.isChecked()
            root = self.ranking_table.invisibleRootItem()
            
            # 모든 아이템의 체크박스 상태 변경
            for i in range(root.childCount()):
                item = root.child(i)
                checkbox_widget = self.ranking_table.itemWidget(item, 0)
                if checkbox_widget and hasattr(checkbox_widget, 'setChecked'):
                    checkbox_widget.setChecked(is_checked)
            
            # 삭제 버튼 상태 업데이트
            self.update_delete_button_state()
            
        except Exception as e:
            logger.error(f"헤더 체크박스 변경 처리 실패: {e}")
    
    def toggle_all_checkboxes(self):
        """헤더 체크박스로 전체 선택/해제"""
        is_checked = self.header_checkbox.isChecked()
        root = self.ranking_table.invisibleRootItem()
        
        for i in range(root.childCount()):
            item = root.child(i)
            checkbox_widget = self.ranking_table.itemWidget(item, 0)
            if checkbox_widget:
                checkbox_widget.setChecked(is_checked)
        
        self.update_delete_button_state()
    
    def update_delete_button_state(self):
        """체크된 항목에 따라 삭제 버튼 활성화/비활성화 (원본과 완전 동일)"""
        checked_count = 0
        root = self.ranking_table.invisibleRootItem()
        
        for i in range(root.childCount()):
            item = root.child(i)
            checkbox_widget = self.ranking_table.itemWidget(item, 0)
            if checkbox_widget and hasattr(checkbox_widget, 'isChecked') and checkbox_widget.isChecked():
                checked_count += 1
        
        # 삭제 버튼 상태 업데이트 (원본과 동일하게 개수 표시)
        if hasattr(self, 'delete_keywords_button'):
            self.delete_keywords_button.setEnabled(checked_count > 0)
            self.delete_keywords_button.setText(f"🗑️ 선택 키워드 삭제 ({checked_count}개)")
    
    def delete_selected_keywords(self):
        """선택된 키워드들 삭제"""
        if not self.current_project:
            return
        
        # 선택된 키워드 수집
        selected_keyword_ids = []
        selected_keywords = []
        root = self.ranking_table.invisibleRootItem()
        
        for i in range(root.childCount()):
            item = root.child(i)
            checkbox_widget = self.ranking_table.itemWidget(item, 0)
            if checkbox_widget and checkbox_widget.isChecked():
                keyword_id = item.data(1, Qt.UserRole)
                keyword_text = item.text(1)
                if keyword_id:
                    selected_keyword_ids.append(keyword_id)
                    selected_keywords.append(keyword_text)
        
        if not selected_keyword_ids:
            return
        
        # 확인 다이얼로그 (화면 중앙에 표시)
        from src.toolbox.ui_kit import ModernConfirmDialog
        
        # 메인 윈도우를 부모로 설정하여 중앙에 표시
        main_window = self.window()
        if ModernConfirmDialog.question(
            main_window,
            "키워드 삭제 확인",
            f"선택한 {len(selected_keywords)}개 키워드를 삭제하시겠습니까?\n\n" +
            "삭제할 키워드:\n" + "\n".join([f"• {kw}" for kw in selected_keywords[:5]]) +
            (f"\n... 외 {len(selected_keywords)-5}개" if len(selected_keywords) > 5 else ""),
            "삭제", "취소"
        ):
            # 키워드 삭제 실행
            success_count = 0
            for keyword_text in selected_keywords:
                try:
                    if ranking_table_view_model.delete_keyword(self.current_project_id, keyword_text):
                        success_count += 1
                except Exception as e:
                    logger.error(f"키워드 삭제 실패: {e}")
            
            if success_count > 0:
                log_manager.add_log(f"✅ {success_count}개 키워드가 삭제되었습니다.", "success")
                # 테이블 새로고침
                self.update_ranking_table(self.current_project_id)
    
    
    
    
    
    def show_progress(self, message: str, show_bar: bool = False):
        """진행 상황 표시"""
        self.progress_frame.setVisible(True)
        self.progress_label.setText(message)
        if show_bar:
            self.progress_bar.setVisible(True)
        else:
            self.progress_bar.setVisible(False)
    
    def hide_progress(self):
        """진행 상황 숨기기"""
        self.progress_frame.setVisible(False)
        self.progress_bar.setVisible(False)
    
    def set_selected_projects(self, selected_projects):
        """다중 선택된 프로젝트들 설정"""
        try:
            self.selected_projects = selected_projects or []
            logger.info(f"선택된 프로젝트 수: {len(self.selected_projects)}")
            
            # 저장 버튼 텍스트 업데이트
            if len(self.selected_projects) > 1:
                self.save_button.setText(f"💾 저장 ({len(self.selected_projects)}개)")
            elif len(self.selected_projects) == 1:
                self.save_button.setText("💾 저장")
            else:
                self.save_button.setText("💾 저장")
                
        except Exception as e:
            logger.error(f"선택된 프로젝트 설정 오류: {e}")
    
    def _export_multiple_projects_to_excel(self, projects, file_path):
        """다중 프로젝트를 엑셀의 여러 시트로 저장"""
        try:
            import openpyxl
            from .excel_export import rank_tracking_excel_exporter
            
            # 새 워크북 생성
            workbook = openpyxl.Workbook()
            
            # 기본 시트 제거
            if 'Sheet' in workbook.sheetnames:
                del workbook['Sheet']
            
            success_count = 0
            
            for i, project in enumerate(projects, 1):
                try:
                    project_id = project.id if hasattr(project, 'id') else project
                    project_name = project.current_name if hasattr(project, 'current_name') else f"프로젝트 {project_id}"
                    
                    # 시트명 생성 (Sheet1, Sheet2 형식)
                    sheet_name = f"Sheet{i}"
                    
                    # 새 시트 생성
                    worksheet = workbook.create_sheet(title=sheet_name)
                    
                    # 임시 파일에 단일 프로젝트 데이터 생성
                    import tempfile
                    import os
                    
                    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as temp_file:
                        temp_path = temp_file.name
                    
                    # 단일 프로젝트 엑셀 생성
                    single_success = rank_tracking_excel_exporter.export_ranking_history_to_excel(
                        project_id, 
                        temp_path
                    )
                    
                    if single_success:
                        # 임시 파일에서 데이터 읽어오기
                        temp_workbook = openpyxl.load_workbook(temp_path)
                        temp_worksheet = temp_workbook.active
                        
                        # 데이터 복사
                        for row in temp_worksheet.iter_rows(values_only=True):
                            worksheet.append(row)
                        
                        # 스타일 복사 (기본적인 스타일만)
                        for row_num, row in enumerate(temp_worksheet.iter_rows(), 1):
                            for col_num, cell in enumerate(row, 1):
                                target_cell = worksheet.cell(row=row_num, column=col_num)
                                if cell.font:
                                    target_cell.font = cell.font
                                if cell.fill:
                                    target_cell.fill = cell.fill
                                if cell.alignment:
                                    target_cell.alignment = cell.alignment
                        
                        temp_workbook.close()
                        success_count += 1
                        logger.info(f"프로젝트 '{project_name}' Sheet{i}에 저장 완료")
                    
                    # 임시 파일 삭제
                    try:
                        os.unlink(temp_path)
                    except:
                        pass
                        
                except Exception as e:
                    logger.error(f"프로젝트 {project_id} 처리 실패: {e}")
                    continue
            
            if success_count > 0:
                # 최종 파일 저장
                workbook.save(file_path)
                workbook.close()
                logger.info(f"다중 프로젝트 엑셀 저장 완료: {success_count}/{len(projects)}개")
                return True
            else:
                workbook.close()
                logger.error("저장할 프로젝트 데이터가 없습니다")
                return False
                
        except Exception as e:
            logger.error(f"다중 프로젝트 엑셀 저장 실패: {e}")
            return False
    
    def export_data(self):
        """순위 이력을 엑셀로 내보내기 (다중 프로젝트 지원)"""
        try:
            from .excel_export import rank_tracking_excel_exporter
            from src.toolbox.ui_kit import ModernSaveCompletionDialog, ModernInfoDialog
            from PySide6.QtWidgets import QFileDialog
            
            # 내보낼 프로젝트 결정
            export_projects = []
            
            if len(self.selected_projects) > 1:
                # 다중 선택된 경우
                export_projects = self.selected_projects
                export_type = "다중 프로젝트"
            elif len(self.selected_projects) == 1:
                # 단일 선택된 경우
                export_projects = self.selected_projects
                export_type = "단일 프로젝트"
            elif self.current_project and self.current_project_id:
                # 현재 프로젝트만 있는 경우
                export_projects = [self.current_project]
                export_type = "현재 프로젝트"
            else:
                log_manager.add_log("⚠️ 내보낼 프로젝트가 선택되지 않았습니다.", "warning")
                return
            
            logger.info(f"내보낼 프로젝트 수: {len(export_projects)}, 타입: {export_type}")
            
            # 디버깅: 프로젝트 정보 확인
            if export_projects:
                first_project = export_projects[0]
                project_id = first_project.id if hasattr(first_project, 'id') else first_project
                logger.info(f"디버깅: 프로젝트 ID = {project_id}, 프로젝트 객체 타입 = {type(first_project)}")
                
                # 키워드가 있는지 확인
                from .service import rank_tracking_service
                keywords = rank_tracking_service.get_project_keywords(project_id)
                logger.info(f"디버깅: 키워드 수 = {len(keywords) if keywords else 0}")
                if keywords:
                    logger.info(f"디버깅: 첫 번째 키워드 = {keywords[0].keyword if keywords[0] else 'None'}")
            
            # 파일명 및 저장 다이얼로그
            if len(export_projects) == 1:
                # 단일 프로젝트
                project_id = export_projects[0].id if hasattr(export_projects[0], 'id') else export_projects[0]
                default_filename = rank_tracking_excel_exporter.get_default_filename(project_id)
                dialog_title = "순위 이력 저장"
            else:
                # 다중 프로젝트
                from datetime import datetime
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                default_filename = f"순위이력_다중프로젝트_{len(export_projects)}개_{timestamp}.xlsx"
                dialog_title = f"다중 프로젝트 순위 이력 저장 ({len(export_projects)}개)"
            
            # 파일 저장 다이얼로그
            file_path, _ = QFileDialog.getSaveFileName(
                self,
                dialog_title,
                default_filename,
                "Excel files (*.xlsx)"
            )
            
            if not file_path:
                return
            
            logger.info(f"디버깅: 저장할 파일 경로 = {file_path}")
            
            # 진행 상황 표시
            self.show_progress("📊 엑셀 파일 생성 중...", show_bar=False)
            
            # 엑셀 파일 생성
            success = False
            result_message = ""
            
            try:
                if len(export_projects) == 1:
                    # 단일 프로젝트 저장
                    project_id = export_projects[0].id if hasattr(export_projects[0], 'id') else export_projects[0]
                    logger.info(f"디버깅: 단일 프로젝트 저장 시작, project_id = {project_id}")
                    
                    success = rank_tracking_excel_exporter.export_ranking_history_to_excel(
                        project_id, 
                        file_path
                    )
                    logger.info(f"디버깅: 엑셀 저장 결과 = {success}")
                    result_message = f"프로젝트: {export_projects[0].current_name if hasattr(export_projects[0], 'current_name') else '알 수 없음'}"
                else:
                    # 다중 프로젝트 저장
                    logger.info(f"디버깅: 다중 프로젝트 저장 시작")
                    success = self._export_multiple_projects_to_excel(export_projects, file_path)
                    result_message = f"{len(export_projects)}개 프로젝트가 별도 시트로 저장되었습니다."
                    
            except Exception as export_error:
                logger.error(f"디버깅: 엑셀 저장 중 예외 발생: {export_error}")
                import traceback
                logger.error(f"디버깅: 스택 트레이스:\n{traceback.format_exc()}")
                success = False
            
            # 진행 상황 숨기기
            self.hide_progress()
            
            if success:
                log_manager.add_log(f"✅ 순위 이력이 엑셀로 저장되었습니다: {file_path}", "success")
                
                # 저장 완료 다이얼로그 (메인 윈도우를 부모로 설정)
                main_window = self.window()  # 최상위 윈도우 찾기
                ModernSaveCompletionDialog.show_save_completion(
                    main_window,
                    "저장 완료",
                    f"순위 이력이 성공적으로 저장되었습니다.\n\n{result_message}",
                    file_path
                )
            else:
                log_manager.add_log("❌ 순위 이력 저장에 실패했습니다.", "error")
                main_window = self.window()  # 최상위 윈도우 찾기
                ModernInfoDialog.warning(main_window, "저장 실패", "순위 이력 저장 중 오류가 발생했습니다.")
                
        except Exception as e:
            self.hide_progress()
            logger.error(f"순위 이력 내보내기 실패: {e}")
            import traceback
            logger.error(f"전체 스택 트레이스:\n{traceback.format_exc()}")
            from src.toolbox.ui_kit import ModernInfoDialog
            ModernInfoDialog.warning(
                self, 
                "내보내기 오류", 
                f"엑셀 내보내기 중 오류가 발생했습니다:\n{str(e)}"
            )
            log_manager.add_log(f"❌ 순위 이력 내보내기 중 오류가 발생했습니다: {str(e)}", "error")
    
    def on_ranking_check_finished(self, project_id, success, message, results):
        """순위 확인 완료 - 프로젝트별 처리"""
        # 워커 정리는 ranking_worker_manager에서 처리
        
        # 현재 보고 있는 프로젝트인 경우에만 UI 업데이트
        if self.current_project_id and self.current_project_id == project_id:
            self.update_button_state_for_current_project(running=False)
            self.update_ranking_table(project_id)
            self.hide_progress()
        
        # 워커에서 이미 상품명으로 로그를 출력하므로 여기서는 중복 로그 제거
        if success:
            pass  # 워커에서 이미 로그 출력
        else:
            log_manager.add_log(f"❌ {self.current_project.current_name if self.current_project else '프로젝트'} 순위 확인 실패: {message}", "error")
        
        logger.info(f"프로젝트 {project_id} 순위 확인 완료: {message}")
    
    def on_progress_updated(self, project_id, current, total):
        """진행률 업데이트 - 프로젝트별 처리"""
        # 현재 보고 있는 프로젝트인 경우에만 UI 업데이트
        if self.current_project_id and self.current_project_id == project_id:
            self.show_progress(f"순위 확인 중... ({current}/{total})", show_bar=True)
            percentage = int((current / total) * 100) if total > 0 else 0
            self.progress_bar.setValue(percentage)
    
    def on_keyword_rank_updated(self, project_id, keyword_id, keyword, rank, volume):
        """키워드 순위 업데이트 - 프로젝트별 처리"""
        logger.info(f"🎯🎯🎯 순위 업데이트 시그널 수신! 프로젝트={project_id}, 키워드ID={keyword_id}, 키워드={keyword}, 순위={rank}")
        
        # 현재 보고 있는 프로젝트인 경우에만 UI 업데이트
        if self.current_project_id and self.current_project_id == project_id:
            logger.info(f"🎯🎯🎯 현재 보고 있는 프로젝트와 일치함. UI 업데이트 실행")
            # 실시간 테이블 업데이트 로직
            self.update_single_keyword_rank(keyword_id, keyword, rank, volume)
        else:
            logger.info(f"🎯🎯🎯 현재 프로젝트 ID({self.current_project_id})와 다름. UI 업데이트 건너뜀")
    
    def add_new_ranking_column_with_time(self, time_str: str):
        """새로운 순위 컬럼을 월검색량 바로 다음(4번째)에 삽입"""
        try:
            logger.info(f"새 순위 컬럼 추가 시작: {time_str}")
            
            # 삽입할 위치 (월검색량 다음 = 4번째 인덱스)
            insert_position = 4
            
            column_count = self.ranking_table.columnCount()
            row_count = self.ranking_table.topLevelItemCount()
            logger.info(f"현재 컬럼 수: {column_count}, 행 수: {row_count}")
            
            # 새 컬럼 추가 (맨 뒤에 임시로 추가)
            self.ranking_table.setColumnCount(column_count + 1)
            
            # 헤더 재배치: 4번째 위치에 새 시간 헤더 삽입
            formatted_time = format_date(time_str)
            
            # 기존 헤더들을 수집하고 4번째 위치에 새 헤더 삽입
            new_headers = []
            header_item = self.ranking_table.headerItem()
            
            for i in range(column_count + 1):  # 새로 추가된 컬럼까지 포함
                if i < insert_position:
                    # 4번째 위치 전까지는 기존 헤더 유지
                    if header_item and i < column_count:
                        new_headers.append(header_item.text(i))
                    else:
                        new_headers.append("")
                elif i == insert_position:
                    # 4번째 위치에 새 시간 헤더 삽입
                    new_headers.append(formatted_time)
                else:
                    # 4번째 위치 이후는 기존 헤더를 한 칸씩 뒤로 이동
                    original_index = i - 1
                    if header_item and original_index < column_count:
                        new_headers.append(header_item.text(original_index))
                    else:
                        new_headers.append("")
            
            # 새 헤더 적용
            self.ranking_table.setHeaderLabels(new_headers)
            
            # 모든 행의 데이터 재배치: 4번째 위치에 "-" 삽입
            total_items = self.ranking_table.topLevelItemCount()
            
            for i in range(total_items):
                try:
                    item = self.ranking_table.topLevelItem(i)
                    if item:
                        keyword_name = item.text(1)  # 키워드명
                        
                        # 기존 데이터를 뒤에서부터 한 칸씩 뒤로 이동
                        for col in range(column_count, insert_position, -1):
                            old_text = item.text(col - 1) if col - 1 < item.columnCount() else ""
                            item.setText(col, old_text)
                        
                        # 4번째 위치에 "-" 삽입
                        item.setText(insert_position, "-")
                except Exception as item_e:
                    logger.error(f"행 {i} 처리 실패: {item_e}")
            
            # UI 강제 업데이트
            self.ranking_table.viewport().update()
            self.ranking_table.header().updateGeometry()  # update() 대신 updateGeometry() 사용
            self.ranking_table.resizeColumnToContents(insert_position)  # 새 컬럼 크기 조정
            QApplication.processEvents()
            
            logger.info(f"4번째 위치에 새 순위 컬럼 '{formatted_time}' 삽입 완료")
            
        except Exception as e:
            logger.error(f"새 순위 컬럼 추가 실패: {e}")
            import traceback
            traceback.print_exc()
    
    def setup_ranking_table_for_new_check(self, project_id: int, current_time: str):
        """순위 확인용 기본 테이블 구성 (키워드만 + 새 시간 컬럼)"""
        try:
            logger.info(f"순위 확인용 테이블 구성: 프로젝트 {project_id}")
            
            # 기본 헤더 설정
            headers = ["", "키워드", "카테고리", "월검색량"]
            
            # 새로운 시간 컬럼 추가
            formatted_time = self.format_date(current_time)
            headers.append(formatted_time)
            
            # 테이블 완전 초기화
            self.ranking_table.clear()
            self.ranking_table.setColumnCount(len(headers))
            self.ranking_table.setHeaderLabels(headers)
            
            # 키워드만 가져와서 테이블 구성 (기존 순위 데이터 무시)
            keywords = ranking_table_view_model.get_project_keywords(project_id)
            
            for keyword in keywords:
                # 리스트로 아이템 데이터 준비
                row_data = ["", keyword.keyword]  # 첫 번째는 체크박스용 빈 문자열
                
                # 카테고리 추가
                category = keyword.category or '-'
                row_data.append(category)
                
                # 월검색량
                monthly_vol = keyword.monthly_volume if keyword.monthly_volume is not None else -1
                if monthly_vol == -1:
                    volume_text = "-"
                elif monthly_vol == 0:
                    volume_text = "0"
                else:
                    volume_text = f"{monthly_vol:,}"
                row_data.append(volume_text)
                
                # 새 시간 컬럼에 "-" 추가
                row_data.append("-")
                
                # QTreeWidgetItem 생성 및 추가
                item = QTreeWidgetItem(row_data)
                item.setData(1, Qt.UserRole, keyword.id)  # 키워드 ID 저장
                
                # 체크박스 설정
                item.setFlags(item.flags() | Qt.ItemIsUserCheckable)
                item.setCheckState(0, Qt.Checked if keyword.is_active else Qt.Unchecked)
                
                self.ranking_table.addTopLevelItem(item)
            
            # 헤더 체크박스 설정
            self.setup_header_checkbox()
            
            logger.info(f"✅ 순위 확인용 테이블 구성 완료: {len(keywords)}개 키워드, 새 컬럼 '{formatted_time}'")
            
        except Exception as e:
            logger.error(f"❌ 순위 확인용 테이블 구성 실패: {e}")
            import traceback
            logger.error(traceback.format_exc())
    
    def update_single_keyword_rank(self, keyword_id, keyword, rank, volume):
        """단일 키워드의 순위를 실시간으로 업데이트"""
        try:
            logger.info(f"실시간 순위 업데이트 요청: 키워드ID={keyword_id}, 키워드={keyword}, 순위={rank}")
            
            # 테이블에서 해당 키워드 찾기
            found = False
            for i in range(self.ranking_table.topLevelItemCount()):
                item = self.ranking_table.topLevelItem(i)
                stored_keyword_id = item.data(1, Qt.UserRole) if item else None
                logger.debug(f"행 {i}: 저장된 키워드ID={stored_keyword_id}, 찾는 키워드ID={keyword_id}")
                
                if item and stored_keyword_id == keyword_id:
                    found = True
                    # 새로 생성한 순위 컬럼(4번째)에 순위 업데이트
                    ranking_column = 4  # 월검색량(3) 다음 위치
                    logger.info(f"키워드 찾음! 업데이트할 컬럼: {ranking_column} (4번째 컬럼)")
                    
                    # 순위 표시
                    rank_display = format_rank_display(rank)
                    item.setText(ranking_column, rank_display)
                    logger.info(f"순위 텍스트 설정 완료: {rank_display}")
                    
                    # 순위에 따른 색상 설정
                    color = get_rank_color(rank, "foreground")
                    item.setForeground(ranking_column, QColor(color))
                    
                    # 정렬용 데이터 설정
                    sort_rank = 201 if (rank == 0 or rank > 200) else rank
                    item.setData(ranking_column, Qt.UserRole, sort_rank)
                    logger.info(f"키워드 {keyword} 실시간 업데이트 완료")
                    break
            
            if not found:
                logger.warning(f"키워드 ID {keyword_id} ('{keyword}')에 해당하는 테이블 행을 찾을 수 없음")
                    
        except Exception as e:
            logger.error(f"키워드 순위 실시간 업데이트 실패: {e}")
    
    def _apply_category_color(self, item, category: str):
        """카테고리 색상 적용 (기본정보에서 카테고리 바로 확인)"""
        if not category or category == "-":
            return
            
        try:
            # 기본정보 화면에서 카테고리 정보 바로 가져오기
            if hasattr(self, 'category_label') and self.category_label:
                project_category = self.category_label.text()
                
                if project_category and project_category != "-":
                    logger.info(f"🔍 카테고리 비교: 프로젝트='{project_category}' vs 키워드='{category}'")
                    color = get_category_match_color(project_category, category)
                    item.setForeground(2, QColor(color))
                    logger.info(f"✅ 색상 적용: {color} ({'초록' if color == '#059669' else '빨강' if color == '#DC2626' else '회색'})")
                else:
                    logger.warning("기본정보의 카테고리가 비어있음")
            else:
                logger.warning("category_label이 없음")
                    
        except Exception as e:
            logger.error(f"카테고리 색상 설정 실패: {e}")
    
    def _update_monthly_volume_display(self, item, monthly_volume: int):
        """월검색량 표시 업데이트 (중복 코드 제거)"""
        formatted_volume = format_monthly_volume(monthly_volume)
        item.setText(3, formatted_volume)
        item.setData(3, Qt.UserRole, monthly_volume)
    
    def _find_keyword_item(self, keyword: str):
        """테이블에서 키워드 아이템 찾기 (중복 코드 제거)"""
        root = self.ranking_table.invisibleRootItem()
        for i in range(root.childCount()):
            item = root.child(i)
            if item and item.text(1) == keyword:
                return item
        return None
    
    # show_keyword_addition_progress, update_keyword_progress 함수들 제거됨
    # 백그라운드 처리에서 기존 progress_frame 사용
    
    # 사용하지 않는 진행률 관련 함수들 제거됨
    # 기존 progress_frame과 progress_bar를 재사용
    
    def add_keywords_to_table_immediately(self, keywords: list):
        """테이블에 키워드 즉시 추가 (월검색량/카테고리는 나중에, 원본과 동일)"""
        try:
            for keyword in keywords:
                # 새 항목 생성 (SortableTreeWidgetItem는 같은 파일에 정의되어 있음)
                item = SortableTreeWidgetItem([])
                
                # 체크박스 생성 (컬럼 0)
                checkbox = self._create_item_checkbox()
                self.ranking_table.addTopLevelItem(item)
                self.ranking_table.setItemWidget(item, 0, checkbox)
                
                # 키워드 (컬럼 1)
                item.setText(1, keyword)
                item.setData(1, Qt.UserRole, keyword)  # 키워드 ID는 나중에 설정
                
                # 카테고리 (컬럼 2) - 일단 "-"로 표시
                item.setText(2, "-")
                item.setData(2, Qt.UserRole, 0)
                
                # 월검색량 (컬럼 3) - 일단 "-"로 표시 (아직 검색하지 않음)
                item.setText(3, "-")
                item.setData(3, Qt.UserRole, -1)
                
                # 순위 컬럼들 (4번 이후) - 모두 "-"로 초기화
                column_count = self.ranking_table.columnCount()
                for col in range(4, column_count):
                    item.setText(col, "-")
                    item.setData(col, Qt.UserRole, 202)  # 정렬 시 맨 아래
                
            # 헤더 체크박스 재설정
            self.reset_header_checkbox()
            
            log_manager.add_log(f"✅ 테이블에 {len(keywords)}개 키워드 추가 완료", "success")
            
        except Exception as e:
            log_manager.add_log(f"❌ 테이블 업데이트 실패: {e}", "error")
    
    # ========== 새로고침 관련 메서드들 ==========
    # 새로고침은 키워드 추가와 동일한 방식(_start_background_keyword_update)을 사용
    
    def _force_refresh_ranking_table(self):
        """테이블 완전 새로고침 (기존 구조 완전 제거 후 재구성)"""
        if not self.current_project_id:
            return
            
        try:
            # 1단계: 기존 헤더 체크박스 제거
            if hasattr(self, 'header_checkbox'):
                try:
                    self.header_checkbox.deleteLater()
                    delattr(self, 'header_checkbox')
                except:
                    pass
            
            # 2단계: 테이블 완전 초기화
            self.ranking_table.clear()
            self.ranking_table.setColumnCount(0)
            self.ranking_table.setHeaderLabels([])
            
            # 3단계: 짧은 대기 시간 후 재구성 (UI 업데이트 보장)
            from PySide6.QtCore import QTimer
            QTimer.singleShot(100, lambda: self._rebuild_ranking_table())
            
        except Exception as e:
            logger.error(f"테이블 새로고침 오류: {e}")
            # 기본 새로고침으로 대체
            self.update_ranking_table(self.current_project_id)
    
    def _rebuild_ranking_table(self):
        """테이블 완전 재구성"""
        if not self.current_project_id:
            return
        
        try:
            # 기본 테이블 업데이트 호출
            self.update_ranking_table(self.current_project_id)
            logger.info("테이블 완전 재구성 완료")
        except Exception as e:
            logger.error(f"테이블 재구성 오류: {e}")
    
    def refresh_project_display(self):
        """프로젝트 표시 새로고침"""
        if self.current_project_id:
            self.update_project_info(self.current_project_id)
    
    def check_rankings(self):
        """순위 확인 - 프로젝트별 독립 실행"""
        logger.info("💥💥💥 순위 확인 버튼 클릭됨!")
        
        if not self.current_project:
            logger.warning("현재 프로젝트가 선택되지 않음")
            return
        
        project_id = self.current_project_id
        logger.info(f"💥 순위 확인 시작: 프로젝트 ID {project_id}, 프로젝트명: {self.current_project.current_name if self.current_project else 'None'}")
        
        # 워커 매니저를 통해 순위 확인 시작
        success = ranking_worker_manager.start_ranking_check(project_id)
        if success:
            # 현재 프로젝트의 버튼 상태 업데이트
            self.update_button_state_for_current_project(running=True)
            
            # 현재 저장된 시간으로 컬럼 추가
            current_time = ranking_worker_manager.get_current_time(project_id)
            self.add_new_ranking_column_with_time(current_time)
            
            # 즉시 진행률 표시 시작
            self.show_progress("순위 확인 준비 중...", show_bar=True)
            self.progress_bar.setValue(0)
            
            logger.info(f"프로젝트 {project_id} 순위 확인 워커 시작 완료")
        else:
            logger.info(f"프로젝트 {project_id}의 순위 확인이 이미 실행 중입니다.")
    
    def update_button_state_for_current_project(self, running=False):
        """현재 프로젝트의 버튼 상태 업데이트"""
        if running:
            self.check_button.setEnabled(False)
            self.check_button.setText("⏳ 확인 중...")
            self.stop_button.setEnabled(True)
        else:
            self.check_button.setEnabled(True)
            self.check_button.setText("🏆 순위 확인")
            self.stop_button.setEnabled(False)
    

    def on_ranking_finished(self, project_id, success, message, results):
        """순위 확인 완료 처리"""
        logger.info(f"프로젝트 {project_id} 순위 확인 완료: {success}")
            
        # 현재 프로젝트인 경우 UI 업데이트
        if project_id == self.current_project_id:
            self.update_button_state_for_current_project(running=False)
            self.hide_progress()
            # 테이블 새로고침하여 완료된 순위 결과 표시
            self.update_ranking_table(project_id)

    def stop_ranking_check(self):
        """순위 확인 정지 - 현재 프로젝트만"""
        if not self.current_project:
            return
            
        project_id = self.current_project_id
        ranking_worker_manager.stop_ranking_check(project_id)
        logger.info(f"프로젝트 {project_id} 순위 확인 정지 요청")
    
    def add_keyword(self):
        """키워드 추가 다이얼로그"""
        if not self.current_project_id:
            from src.toolbox.ui_kit import ModernInfoDialog
            ModernInfoDialog.warning(
                self, 
                "프로젝트 선택 필요", 
                "📋 기존 프로젝트에 추가하려면: 왼쪽 목록에서 프로젝트를 클릭하세요\n\n" +
                "➕ 새 프로젝트를 만들려면: \"새 프로젝트\" 버튼을 클릭하세요"
            )
            return
        
        from .keyword_dialogs import AddKeywordsDialog
        dialog = AddKeywordsDialog(self.current_project, self)
        
        if dialog.exec() == QDialog.Accepted:
            # 키워드 가져오기
            keywords = dialog.get_keywords()
            if keywords:
                # 즉시 키워드 추가 시작
                
                added_keywords = []
                duplicate_keywords = []
                
                # 1단계: 즉시 DB에 키워드 추가 (빠른 반응)
                for keyword in keywords:
                    try:
                        # ViewModel을 통해 키워드 추가 (이력 기록 포함)
                        keyword_obj = ranking_table_view_model.add_keyword(self.current_project_id, keyword)
                        if keyword_obj:  # 성공적으로 추가된 경우
                            added_keywords.append(keyword)
                            log_manager.add_log(f"✅ '{keyword}' 키워드 추가 완료", "success")
                        else:
                            duplicate_keywords.append(keyword)
                            log_manager.add_log(f"⚠️ '{keyword}' 키워드는 중복입니다.", "warning")
                    except Exception as e:
                        if "이미 등록되어 있습니다" in str(e):
                            duplicate_keywords.append(keyword)
                            log_manager.add_log(f"⚠️ '{keyword}' 키워드는 중복입니다.", "warning")
                        else:
                            logger.error(f"키워드 추가 실패: {keyword}, 오류: {e}")
                            log_manager.add_log(f"❌ '{keyword}' 키워드 추가 실패: {e}", "error")
                
                if added_keywords:
                    # 즉시 테이블 완전 새로고침 (다른 프로젝트 갔다 올 때처럼)
                    self.update_ranking_table(self.current_project_id)
                    log_manager.add_log(f"🎉 {len(added_keywords)}개 키워드 추가 완료!", "success")
                    
                    # 월검색량/카테고리 백그라운드 업데이트 시작
                    self._start_background_keyword_update(added_keywords)
                    
                    if duplicate_keywords:
                        log_manager.add_log(f"⚠️ {len(duplicate_keywords)}개 키워드는 중복으로 건너뜀", "warning")
                else:
                    # 추가된 키워드가 없을 때는 특별한 처리 없음
                    pass
    
    def _start_background_keyword_update(self, keywords):
        """백그라운드 키워드 정보 업데이트 (프로젝트 독립적)"""
        if not keywords or not self.current_project_id:
            return
        
        # 키워드 정보 워커 매니저를 통해 백그라운드 작업 시작
        success = keyword_info_worker_manager.start_keyword_info_update(
            self.current_project_id, 
            keywords, 
            self.current_project
        )
        
        if success:
            log_manager.add_log(f"🔍 '{', '.join(keywords)}' 키워드 월검색량/카테고리 조회를 시작합니다.", "info")
        else:
            log_manager.add_log(f"⚠️ 키워드 정보 조회 시작 실패", "warning")
    
    # 키워드 정보 워커 매니저 시그널 핸들러들
    def on_keyword_info_progress_updated(self, project_id: int, current: int, total: int, current_keyword: str):
        """키워드 정보 업데이트 진행률 처리"""
        # 현재 프로젝트의 진행률만 표시
        if project_id == self.current_project_id:
            if hasattr(self, 'progress_bar') and hasattr(self, 'progress_label'):
                self.progress_bar.setMaximum(total)
                self.progress_bar.setValue(current)
                self.progress_label.setText(f"🔍 월검색량/카테고리 조회 중... ({current}/{total}) - {current_keyword}")
                self.progress_frame.setVisible(True)
                self.progress_bar.setVisible(True)
    
    def on_keyword_category_updated(self, project_id: int, keyword: str, category: str):
        """키워드 카테고리 업데이트 처리"""
        # 현재 프로젝트의 카테고리만 업데이트
        if project_id == self.current_project_id:
            self._update_keyword_category_in_table(keyword, category)
    
    def on_keyword_volume_updated(self, project_id: int, keyword: str, volume: int):
        """키워드 월검색량 업데이트 처리"""
        # 현재 프로젝트의 월검색량만 업데이트
        if project_id == self.current_project_id:
            self._update_keyword_volume_in_table(keyword, volume)
    
    def on_keyword_info_finished(self, project_id: int, success: bool, message: str):
        """키워드 정보 업데이트 완료 처리"""
        # 현재 프로젝트의 완료만 처리
        if project_id == self.current_project_id:
            self.hide_progress()
        
        # 로그는 해당 프로젝트 이름으로 표시
        try:
            from .service import rank_tracking_service
            project = rank_tracking_service.get_project_by_id(project_id)
            project_name = project.current_name if project else f"프로젝트 ID {project_id}"
            
            if success:
                log_manager.add_log(f"✅ {project_name} - {message}", "success")
            else:
                log_manager.add_log(f"❌ {project_name} - {message}", "error")
        except Exception as e:
            logger.error(f"키워드 정보 완료 로그 처리 실패: {e}")
    
    def _update_keyword_category_in_table(self, keyword: str, category: str):
        """테이블에서 키워드 카테고리만 업데이트"""
        try:
            item = self._find_keyword_item(keyword)
            if not item:
                return
            
            # 카테고리 업데이트
            item.setText(2, category or '-')
            
            # 카테고리 색상 적용
            if category != '-' and hasattr(self, 'category_label') and self.category_label:
                project_category = self.category_label.text()
                if project_category and project_category != "-":
                    from .adapters import get_category_match_color
                    from PySide6.QtGui import QColor
                    # 키워드 카테고리에서 괄호 앞 부분만 추출
                    keyword_category_clean = category.split('(')[0].strip()
                    # 프로젝트 카테고리에서 마지막 부분만 추출
                    project_category_base = project_category.split(' > ')[-1] if ' > ' in project_category else project_category
                    color = get_category_match_color(project_category_base, keyword_category_clean)
                    item.setForeground(2, QColor(color))
            
            # 테이블 즉시 새로고침
            self.ranking_table.viewport().update()
            
        except Exception as e:
            logger.error(f"키워드 '{keyword}' 카테고리 테이블 업데이트 실패: {e}")
    
    def _update_keyword_volume_in_table(self, keyword: str, volume: int):
        """테이블에서 키워드 월검색량만 업데이트"""
        try:
            item = self._find_keyword_item(keyword)
            if not item:
                return
            
            # 월검색량 업데이트
            if volume >= 0:
                from .adapters import format_monthly_volume
                volume_text = format_monthly_volume(volume)
                item.setText(3, volume_text)
                item.setData(3, Qt.UserRole, volume)
            else:
                item.setText(3, "-")
                item.setData(3, Qt.UserRole, -1)
            
            # 테이블 즉시 새로고침
            self.ranking_table.viewport().update()
            
        except Exception as e:
            logger.error(f"키워드 '{keyword}' 월검색량 테이블 업데이트 실패: {e}")
    
    def setup_buttons(self, layout):
        """하단 버튼들 설정"""
        # 하단 버튼 영역
        button_layout = QHBoxLayout()
        button_layout.setContentsMargins(0, 10, 0, 0)
        button_layout.setSpacing(10)
        
        # 키워드 추가 버튼 (새 프로젝트 생성과 동일한 색상)
        self.add_keyword_button = QPushButton("🔤 키워드 추가")
        self.add_keyword_button.clicked.connect(self.add_keyword)
        self.add_keyword_button.setEnabled(False)  # 프로젝트 선택 시에만 활성화
        self.add_keyword_button.setStyleSheet(f"""
            QPushButton {{
                background-color: {ModernStyle.COLORS['primary']};
                color: white;
                border: none;
                padding: 10px 15px;
                border-radius: 6px;
                font-weight: 600;
                min-width: 120px;
            }}
            QPushButton:hover {{
                background-color: {ModernStyle.COLORS['primary_hover']};
            }}
            QPushButton:disabled {{
                background-color: {ModernStyle.COLORS['border']};
                color: white;
            }}
        """)
        button_layout.addWidget(self.add_keyword_button)
        
        # 순위 확인 버튼 (키워드 검색기 클리어 버튼과 동일한 warning 색상)
        self.check_button = QPushButton("🔍 순위 확인")
        self.check_button.clicked.connect(self.check_rankings)
        self.check_button.setEnabled(False)  # 프로젝트 선택 시에만 활성화
        self.check_button.setStyleSheet(f"""
            QPushButton {{
                background-color: {ModernStyle.COLORS['warning']};
                color: white;
                border: none;
                padding: 10px 15px;
                border-radius: 6px;
                font-weight: 600;
                min-width: 120px;
            }}
            QPushButton:hover {{
                background-color: #d97706;
            }}
            QPushButton:disabled {{
                background-color: {ModernStyle.COLORS['border']};
                color: white;
            }}
        """)
        button_layout.addWidget(self.check_button)
        
        # 정지 버튼 (기존 스타일)
        self.stop_button = QPushButton("⏹️ 정지")
        self.stop_button.clicked.connect(self.stop_ranking_check)
        self.stop_button.setEnabled(False)
        self.stop_button.setStyleSheet(f"""
            QPushButton {{
                background-color: #dc3545;
                color: white;
                border: none;
                padding: 10px 15px;
                border-radius: 6px;
                font-weight: 600;
                min-width: 80px;
            }}
            QPushButton:hover {{
                background-color: #c82333;
            }}
            QPushButton:disabled {{
                background-color: {ModernStyle.COLORS['border']};
                color: white;
            }}
        """)
        button_layout.addWidget(self.stop_button)
        
        # 오른쪽 끝으로 밀기 위한 스트레치
        button_layout.addStretch()
        
        # 저장 버튼 (공용 success 색상, 오른쪽 끝)
        self.save_button = QPushButton("💾 저장")
        self.save_button.clicked.connect(self.export_data)
        self.save_button.setEnabled(False)  # 프로젝트 선택 시에만 활성화
        self.save_button.setStyleSheet(f"""
            QPushButton {{
                background-color: {ModernStyle.COLORS['success']};
                color: white;
                border: none;
                padding: 10px 15px;
                border-radius: 6px;
                font-weight: 600;
                min-width: 100px;
            }}
            QPushButton:hover {{
                background-color: #059669;
            }}
            QPushButton:disabled {{
                background-color: {ModernStyle.COLORS['border']};
                color: white;
            }}
        """)
        button_layout.addWidget(self.save_button)
        
        layout.addLayout(button_layout)
    
    def refresh_product_info(self):
        """상품 정보 새로고침 - service layer로 위임"""
        if not self.current_project or not self.current_project_id:
            log_manager.add_log("⚠️ 새로고침할 프로젝트가 선택되지 않았습니다.", "warning")
            return
        
        # 버튼 비활성화 (새로고침 중)
        self.refresh_product_button.setEnabled(False)
        self.refresh_product_button.setText("⏳")
        
        try:
            # ViewModel을 통한 새로고침 처리
            success, message = ranking_table_view_model.refresh_project_info(self.current_project_id)
            result = {'success': success, 'message': message}
            
            if result['success']:
                # 프로젝트 기본 정보 다시 로드
                if hasattr(self, 'update_project_info'):
                    self.update_project_info(self.current_project_id)
                
                # 프로젝트 정보 다시 불러오기
                from .service import rank_tracking_service
                updated_project = rank_tracking_service.get_project_by_id(self.current_project_id)
                if updated_project:
                    self.current_project = updated_project
                    
                    # 기본정보 화면 업데이트
                    if hasattr(self, 'product_id_label'):
                        self.product_id_label.setText(updated_project.product_id)
                    if hasattr(self, 'product_name_label'):
                        self.product_name_label.setText(updated_project.current_name)
                    if hasattr(self, 'price_label'):
                        self.price_label.setText(f"{updated_project.price:,}원" if updated_project.price else "-")
                    if hasattr(self, 'category_label'):
                        self.category_label.setText(updated_project.category or "-")
                    if hasattr(self, 'store_name_label'):
                        self.store_name_label.setText(updated_project.store_name or "-")
                
                # 키워드 월검색량 및 카테고리 새로고침 (키워드 추가와 동일한 방식)
                keywords = ranking_table_view_model.get_project_keywords(self.current_project_id)
                if keywords:
                    keyword_names = [kw.keyword for kw in keywords]
                    log_manager.add_log(f"🔄 {len(keyword_names)}개 키워드의 월검색량/카테고리를 업데이트합니다...", "info")
                    # 키워드 추가와 동일한 방식으로 백그라운드 업데이트 (실시간 색상 적용 포함)
                    self._start_background_keyword_update(keyword_names)
                else:
                    log_manager.add_log("📝 새로고침할 키워드가 없습니다.", "info")
                
                log_manager.add_log(result['message'], "success")
            else:
                log_manager.add_log(result['message'], "error")
                
        except Exception as e:
            logger.error(f"상품 정보 새로고침 실패: {e}")
            log_manager.add_log(f"❌ 상품 정보 새로고침 중 오류가 발생했습니다: {str(e)}", "error")
        
        finally:
            # 버튼 복원
            self.refresh_product_button.setEnabled(True)
            self.refresh_product_button.setText("⟲")
    
    def show_project_changes(self):
        """프로젝트 변경사항 다이얼로그 표시 - 새로운 ProjectHistoryDialog 사용"""
        if not self.current_project or not self.current_project_id:
            log_manager.add_log("⚠️ 변경사항을 확인할 프로젝트가 선택되지 않았습니다.", "warning")
            return
        
        try:
            # 테이블 헤더에서 가장 최근 시간과 이전 시간 가져오기
            header_times = self.get_header_times()
            current_time = header_times[0] if len(header_times) > 0 else None
            previous_time = header_times[1] if len(header_times) > 1 else None
            
            # 새로운 ProjectHistoryDialog 사용
            from .history_dialog import ProjectHistoryDialog
            dialog = ProjectHistoryDialog(
                self.current_project_id, 
                self.current_project.current_name, 
                self, 
                current_time,
                previous_time
            )
            dialog.exec()
            
        except Exception as e:
            logger.error(f"프로젝트 변경사항 다이얼로그 표시 실패: {e}")
            log_manager.add_log(f"❌ 변경사항 다이얼로그 표시 중 오류가 발생했습니다: {str(e)}", "error")
    
    def get_latest_check_time(self):
        """DB에서 가장 최근 순위 확인 시간 가져오기"""
        try:
            if not self.current_project_id:
                return None
            
            # Foundation DB를 통해 가장 최근 순위 확인 시간 조회
            from src.foundation.db import get_db
            
            db = get_db()
            latest_rankings = db.get_latest_rankings(self.current_project_id, limit=1)
            
            if latest_rankings:
                latest_time = latest_rankings[0].get('created_at')
                if latest_time:
                    # 날짜 포맷팅 - "2025-08-15 22:17:32" 형태로 반환
                    from datetime import datetime
                    if isinstance(latest_time, str):
                        try:
                            dt = datetime.fromisoformat(latest_time.replace('Z', '+00:00'))
                            return dt.strftime("%Y-%m-%d %H:%M:%S")
                        except:
                            return str(latest_time)
                    elif isinstance(latest_time, datetime):
                        return latest_time.strftime("%Y-%m-%d %H:%M:%S")
                    else:
                        return str(latest_time)
            
            return None
            
        except Exception as e:
            logger.error(f"최신 확인 시간 조회 실패: {e}")
            return None
    
    def get_header_times(self):
        """테이블 헤더에서 시간 정보 가져오기"""
        try:
            header_times = []
            header_item = self.ranking_table.headerItem()
            
            if header_item:
                for col in range(4, self.ranking_table.columnCount()):  # 4번째 컬럼부터 날짜
                    header_text = header_item.text(col)
                    if header_text and "/" in header_text:
                        header_times.append(header_text.strip())
            
            return header_times
            
        except Exception as e:
            logger.error(f"헤더 시간 정보 가져오기 실패: {e}")
            return []
    
    def _show_project_changes_dialog(self, changes_data):
        """변경사항 다이얼로그 UI 표시 (기존 통합관리프로그램 스타일)"""
        from PySide6.QtWidgets import QDialog, QVBoxLayout, QTextEdit, QDialogButtonBox
        
        dialog = QDialog(self)
        dialog.setWindowTitle(f"📝 {self.current_project.current_name} - 변경사항")
        dialog.setMinimumSize(600, 400)
        dialog.setModal(True)
        
        layout = QVBoxLayout()
        
        # 헤더
        header_label = QLabel(f"📊 {self.current_project.current_name}")
        header_label.setStyleSheet(f"""
            QLabel {{
                font-size: 16px;
                font-weight: bold;
                color: {ModernStyle.COLORS['text_primary']};
                margin-bottom: 10px;
                padding: 10px;
                background-color: {ModernStyle.COLORS['bg_secondary']};
                border-radius: 6px;
            }}
        """)
        layout.addWidget(header_label)
        
        # 변경사항 텍스트
        changes_text = QTextEdit()
        changes_text.setReadOnly(True)
        
        if changes_data and changes_data.get('changes'):
            # 변경사항이 있는 경우
            changes_list = changes_data['changes']
            text_content = f"🔄 총 {len(changes_list)}건의 변경사항이 있습니다.\n\n"
            
            for i, change in enumerate(changes_list, 1):
                text_content += f"{i}. {change.get('date', 'N/A')} - {change.get('description', '알 수 없는 변경')}\n"
                if change.get('details'):
                    text_content += f"   상세: {change['details']}\n"
                text_content += "\n"
        else:
            # 변경사항이 없는 경우
            text_content = "📝 아직 기록된 변경사항이 없습니다.\n\n" + \
                          "상품 정보나 키워드가 변경되면 이곳에 기록됩니다."
        
        changes_text.setPlainText(text_content)
        changes_text.setStyleSheet(f"""
            QTextEdit {{
                border: 1px solid {ModernStyle.COLORS['border']};
                border-radius: 6px;
                padding: 10px;
                font-family: 'Consolas', 'Courier New', monospace;
                font-size: 12px;
                background-color: {ModernStyle.COLORS['bg_card']};
                color: {ModernStyle.COLORS['text_primary']};
            }}
        """)
        layout.addWidget(changes_text)
        
        # 버튼
        buttons = QDialogButtonBox(QDialogButtonBox.Ok)
        buttons.accepted.connect(dialog.accept)
        buttons.setStyleSheet(f"""
            QPushButton {{
                background-color: {ModernStyle.COLORS['primary']};
                color: white;
                border: none;
                padding: 8px 16px;
                border-radius: 6px;
                font-size: 13px;
                font-weight: 600;
            }}
            QPushButton:hover {{
                background-color: {ModernStyle.COLORS['primary_hover']};
            }}
        """)
        layout.addWidget(buttons)
        
        dialog.setLayout(layout)
        dialog.exec()
    
    def set_selected_projects(self, selected_projects):
        """선택된 프로젝트들 설정 (다중 선택 지원)"""
        try:
            self.selected_projects = selected_projects or []
            logger.info(f"선택된 프로젝트 수: {len(self.selected_projects)}")
            
            # 저장 버튼 텍스트 업데이트
            if len(self.selected_projects) > 1:
                self.save_button.setText(f"💾 저장 ({len(self.selected_projects)}개)")
            else:
                self.save_button.setText("💾 저장")
                
        except Exception as e:
            logger.error(f"선택된 프로젝트 설정 오류: {e}")
            self.selected_projects = []
    
    def export_data(self):
        """순위 이력 데이터 Excel로 내보내기 (단일/다중 프로젝트 지원)"""
        try:
            # 선택된 프로젝트 확인
            if len(self.selected_projects) > 1:
                # 다중 프로젝트 내보내기
                self._export_multiple_projects()
            elif self.current_project_id:
                # 단일 프로젝트 내보내기
                self._export_single_project()
            else:
                log_manager.add_log("⚠️ 내보낼 프로젝트가 선택되지 않았습니다.", "warning")
        except Exception as e:
            logger.error(f"데이터 내보내기 오류: {e}")
            log_manager.add_log(f"❌ 데이터 내보내기 중 오류가 발생했습니다: {str(e)}", "error")
    
    def _export_single_project(self):
        """단일 프로젝트 Excel 내보내기"""
        try:
            from PySide6.QtWidgets import QFileDialog
            
            # 기본 파일명 생성
            default_filename = rank_tracking_excel_exporter.get_default_filename(self.current_project_id)
            
            # 파일 저장 다이얼로그 (toolbox 대신 기본 사용)
            file_path, _ = QFileDialog.getSaveFileName(
                self, 
                "순위 이력 Excel 저장", 
                default_filename,
                "Excel 파일 (*.xlsx);;모든 파일 (*)"
            )
            
            if file_path:
                success = rank_tracking_excel_exporter.export_ranking_history_to_excel(
                    self.current_project_id, file_path
                )
                if success:
                    log_manager.add_log(f"✅ 순위 이력 Excel 파일이 저장되었습니다: {file_path}", "success")
                    # 모던 저장 완료 다이얼로그 (메인 윈도우를 부모로 설정)
                    try:
                        main_window = self.window()  # 최상위 윈도우 찾기
                        ModernSaveCompletionDialog.show_save_completion(
                            main_window,
                            "저장 완료",
                            f"순위 이력이 성공적으로 저장되었습니다.\n\n프로젝트: {self.current_project.current_name}",
                            file_path
                        )
                    except:
                        self._show_open_folder_dialog(file_path)
                else:
                    log_manager.add_log("❌ 순위 이력 저장에 실패했습니다.", "error")
        except Exception as e:
            logger.error(f"단일 프로젝트 내보내기 오류: {e}")
            log_manager.add_log(f"❌ Excel 내보내기 중 오류가 발생했습니다: {str(e)}", "error")
    
    def _export_multiple_projects(self):
        """다중 프로젝트 Excel 내보내기"""
        try:
            from PySide6.QtWidgets import QFileDialog
            from datetime import datetime
            
            # 프로젝트 ID 리스트 생성 (StyleProxy 오류 방지)
            project_ids = []
            for project in self.selected_projects:
                try:
                    # 프로젝트 객체에서 ID 추출
                    if hasattr(project, 'id'):
                        project_ids.append(project.id)
                    elif isinstance(project, dict) and 'id' in project:
                        project_ids.append(project['id'])
                    else:
                        logger.error(f"프로젝트 {project} 처리 실패: unhashable type: '{type(project).__name__}'")
                        continue
                except Exception as e:
                    logger.error(f"프로젝트 {project} 처리 실패: {e}")
                    continue
            
            if not project_ids:
                log_manager.add_log("❌ 저장할 프로젝트 데이터가 없습니다", "error")
                return
            
            # 기본 파일명 생성
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            default_filename = f"다중프로젝트_순위이력_{len(project_ids)}개_{timestamp}.xlsx"
            
            # 파일 저장 다이얼로그
            file_path, _ = QFileDialog.getSaveFileName(
                self, 
                "다중 프로젝트 순위 이력 Excel 저장", 
                default_filename,
                "Excel 파일 (*.xlsx);;모든 파일 (*)"
            )
            
            if file_path:
                log_manager.add_log(f"🔄 {len(project_ids)}개 프로젝트 Excel 생성 중...", "info")
                
                success = rank_tracking_excel_exporter.export_multiple_projects_to_excel(
                    project_ids, file_path
                )
                if success:
                    log_manager.add_log(f"✅ 다중 프로젝트 순위 이력 Excel 파일이 저장되었습니다: {file_path}", "success")
                    log_manager.add_log(f"📊 총 {len(project_ids)}개 프로젝트가 각각 별도 시트로 저장됨", "info")
                    # 모던 저장 완료 다이얼로그 (메인 윈도우를 부모로 설정)
                    try:
                        main_window = self.window()  # 최상위 윈도우 찾기
                        ModernSaveCompletionDialog.show_save_completion(
                            main_window,
                            "저장 완료",
                            f"다중 프로젝트 순위 이력이 성공적으로 저장되었습니다.\n\n총 {len(project_ids)}개 프로젝트가 각각 별도 시트로 저장되었습니다.",
                            file_path
                        )
                    except:
                        self._show_open_folder_dialog(file_path)
                else:
                    log_manager.add_log("❌ 다중 프로젝트 순위 이력 저장에 실패했습니다.", "error")
                    
        except Exception as e:
            logger.error(f"다중 프로젝트 내보내기 오류: {e}")
            log_manager.add_log(f"❌ 다중 프로젝트 Excel 내보내기 중 오류가 발생했습니다: {str(e)}", "error")
    
    def _show_open_folder_dialog(self, file_path: str):
        """저장 완료 후 폴더 열기 다이얼로그"""
        try:
            from PySide6.QtWidgets import QMessageBox
            import os
            
            # 폴더 열기 확인 다이얼로그 (기본 시스템 다이얼로그 사용)
            reply = QMessageBox.question(
                self,
                "저장 완료",
                f"Excel 파일이 저장되었습니다.\n\n파일 위치를 여시겠습니까?\n\n{file_path}",
                QMessageBox.Yes | QMessageBox.No,
                QMessageBox.Yes
            )
            
            reply = (reply == QMessageBox.Yes)
            
            if reply:
                # 파일 탐색기에서 파일 위치 열기
                import subprocess
                import platform
                
                if platform.system() == "Windows":
                    subprocess.run(['explorer', '/select,', file_path], check=False)
                elif platform.system() == "Darwin":  # macOS
                    subprocess.run(['open', '-R', file_path], check=False)
                else:  # Linux
                    subprocess.run(['xdg-open', os.path.dirname(file_path)], check=False)
                    
        except Exception as e:
            logger.error(f"폴더 열기 다이얼로그 오류: {e}")


