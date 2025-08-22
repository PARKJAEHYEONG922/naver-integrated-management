"""
순위 추적 어댑터 - vendors 레이어 응답을 features 데이터로 가공
Raw API 응답을 비즈니스 로직에서 사용할 수 있는 형태로 변환
엑셀 내보내기 포함
"""
from typing import Optional, Dict, Any, List, TypedDict, Tuple
import re
from datetime import datetime

from src.vendors.naver.developer.shopping_client import shopping_client as naver_shopping
from src.vendors.naver.client_factory import get_keyword_tool_client
from src.toolbox.text_utils import validate_naver_url, extract_product_id, validate_product_id
from src.toolbox.formatters import format_monthly_volume, format_rank, format_datetime
from src.foundation.logging import get_logger
from .models import RANK_OUT_OF_RANGE

logger = get_logger("features.rank_tracking.adapters")


class ProductInfoDTO(TypedDict, total=False):
    """상품 정보 DTO"""
    product_id: str
    name: str
    price: int
    category: str
    store_name: str
    description: str
    image_url: str
    url: str


class RankingCheckDTO(TypedDict, total=False):
    """순위 확인 결과 DTO"""
    success: bool
    rank: int
    total_results: int
    error: str
    keyword: str
    product_id: str


def format_date(date_str: str) -> str:
    """날짜 형식 변환 (8/6 14:26)"""
    dt = _to_dt(date_str)
    return dt.strftime("%m/%d %H:%M") if dt else date_str


def format_date_with_time(date_str: str) -> str:
    """날짜 시간 형식 변환 (2025-08-07 15:23)"""
    dt = _to_dt(date_str)
    return dt.strftime("%Y-%m-%d %H:%M") if dt else date_str


def format_rank_display(rank: int) -> str:
    """순위 숫자를 사용자 친화적인 형태로 포맷팅 (UI 표시용)"""
    if not isinstance(rank, int):
        return "-"
    if rank == RANK_OUT_OF_RANGE or rank > 200:
        return "200위밖"
    elif rank >= 1:
        return f"{rank}위"
    return "-"


def _to_dt(date_str: str):
    """문자열을 datetime 객체로 변환 (DATE와 ISO 형식 모두 처리)"""
    try:
        # 'YYYY-MM-DD HH:MM:SS' 또는 'YYYY-MM-DDTHH:MM:SS[Z]' 형식
        return datetime.fromisoformat(date_str.replace('Z', '+00:00'))
    except:
        try:
            # 'YYYY-MM-DD' 형식 (repository에서 반환)
            return datetime.strptime(date_str, "%Y-%m-%d")
        except:
            return None




def get_rank_color(rank: int, color_type: str = "background") -> str:
    """순위에 따른 색상 반환"""
    if color_type == "background":
        # 배경색 (연한 톤)
        if rank <= 10:
            return "#e8f5e8"  # 연한 초록
        elif rank <= 50:
            return "#fff3cd"  # 연한 노랑
        else:
            return "#f8d7da"  # 연한 빨강
    else:  # foreground/text color
        # 텍스트 색상 (진한 톤)
        if rank == -1 or rank == 0:  # 검색량 없음/API 실패
            return "#6B7280"  # 회색
        elif rank <= 10:
            return "#059669"  # 초록색 (상위 10위)
        elif rank <= 50:
            return "#D97706"  # 주황색 (50위 이내)
        else:
            return "#DC2626"  # 빨간색 (50위 초과)


# 기존 format_monthly_volume은 삭제됨 - toolbox.formatters.format_monthly_volume 사용


def get_category_match_color(project_category: str, keyword_category: str) -> str:
    """카테고리 매칭 결과에 따른 색상 반환"""
    if not project_category or not keyword_category:
        return "#6B7280"  # 회색 (데이터 없음)
    
    # 카테고리 비교를 위한 기본 형태로 변환
    project_base = project_category.split(' > ')[-1] if ' > ' in project_category else project_category
    keyword_base = keyword_category.split(' > ')[-1] if ' > ' in keyword_category else keyword_category
    
    if project_base == keyword_base:
        return "#059669"  # 초록색 (일치)
    else:
        return "#DC2626"  # 빨간색 (불일치)


def clean_product_name(name: str) -> str:
    """상품명 정리 (공백 정규화) - service에서 사용 가능"""
    if not name:
        return ""
    
    import re
    # 연속된 공백을 하나로 통일 및 앞뒤 공백 제거
    clean_name = re.sub(r'\s+', ' ', name).strip()
    
    return clean_name


def smart_product_search(product_name: str, product_id: str) -> Optional[Dict[str, Any]]:
    """상품 스마트 검색 - service에서 사용 가능"""
    try:
        return naver_shopping.smart_product_search(product_name, product_id)
    except Exception as e:
        logger.error(f"스마트 상품 검색 실패: {product_name} ({product_id}): {e}")
        return None


class RankTrackingAdapter:
    """순위 추적 어댑터"""
    
    def __init__(self):
        self.shopping_client = naver_shopping
        self.keyword_client = get_keyword_tool_client()
    
    def extract_product_id_from_url(self, url: str) -> str:
        """네이버 쇼핑 URL에서 상품 ID 추출 (validators 사용)"""
        if not url or not isinstance(url, str):
            raise ValueError("URL이 비어있거나 올바르지 않습니다")
        
        if not validate_naver_url(url):
            raise ValueError(
                f"지원되지 않는 네이버 쇼핑 URL 형식입니다: {url}\n"
                "올바른 형식: https://shopping.naver.com/catalog/XXXXX 또는 "
                "https://smartstore.naver.com/store/products/XXXXX"
            )
        
        product_id = extract_product_id(url)
        if not product_id:
            raise ValueError(f"URL에서 상품 ID를 찾을 수 없습니다: {url}")
        
        if not validate_product_id(product_id):
            raise ValueError(f"유효하지 않은 상품 ID입니다: {product_id}")
        
        return product_id
    
    def get_product_info(self, product_name: str, product_id: str) -> Optional[ProductInfoDTO]:
        """상품 정보 조회 (vendors -> ProductInfoDTO 변환)"""
        try:
            raw_data = self.shopping_client.smart_product_search(product_name, product_id)
            if not raw_data:
                return None
            
            return ProductInfoDTO(
                product_id=raw_data.get('product_id', ''),
                name=self._clean_product_name(raw_data.get('name', '')),
                price=raw_data.get('price', 0),
                category=raw_data.get('category', ''),  # 전체 카테고리 경로 유지
                store_name=raw_data.get('store_name', ''),
                description=raw_data.get('description', ''),
                image_url=raw_data.get('image_url', ''),
                url=raw_data.get('url', '')
            )
            
        except Exception as e:
            logger.error(f"상품 정보 조회 실패: {product_name} ({product_id}): {e}")
            return None
    
    def check_product_rank(self, keyword: str, product_id: str) -> RankingCheckDTO:
        """키워드에서 상품 순위 확인"""
        try:
            rank = self.shopping_client.find_product_rank(keyword, product_id, max_pages=10)
            
            result = RankingCheckDTO(
                success=True,
                rank=rank if rank is not None else RANK_OUT_OF_RANGE,
                total_results=max(rank, 100) if rank is not None else 1000,
                keyword=keyword,
                product_id=product_id
            )
            
            logger.info(f"순위 확인 성공: {keyword} -> {product_id} = {rank or '200+'}위")
            return result
            
        except Exception as e:
            logger.error(f"순위 확인 실패: {keyword} -> {product_id}: {e}")
            return RankingCheckDTO(
                success=False,
                rank=RANK_OUT_OF_RANGE,
                total_results=0,
                error=str(e),
                keyword=keyword,
                product_id=product_id
            )
    
    def get_keyword_monthly_volume(self, keyword: str) -> Optional[int]:
        """키워드 월 검색량 조회 (검색광고 API 활용)"""
        try:
            volume = self.keyword_client.get_single_search_volume(keyword)
            logger.debug(f"월검색량 조회: {keyword} -> {volume}")
            return volume
        except Exception as e:
            logger.warning(f"월검색량 조회 실패: {keyword}: {e}")
            return None
    
    def get_keyword_category(self, keyword: str, sample_size: int = None) -> Optional[str]:
        """키워드 대표 카테고리 조회 (쇼핑 API 활용)"""
        try:
            if sample_size is None:
                from .models import DEFAULT_SAMPLE_SIZE
                sample_size = DEFAULT_SAMPLE_SIZE
            category = self.shopping_client.get_keyword_category(keyword, sample_size=sample_size)
            logger.debug(f"카테고리 조회: {keyword} -> {category}")
            return category
        except Exception as e:
            logger.warning(f"카테고리 조회 실패: {keyword}: {e}")
            return None
    
    def analyze_keyword_for_tracking(self, keyword: str) -> Dict[str, Any]:
        """추적용 키워드 종합 분석 (월검색량 + 카테고리)"""
        result = {
            'keyword': keyword,
            'monthly_volume': -1,
            'category': '',
            'success': False,
            'error_message': None
        }
        
        try:
            # 월검색량 조회
            monthly_volume = self.get_keyword_monthly_volume(keyword)
            if monthly_volume is not None:
                result['monthly_volume'] = monthly_volume
            
            # 카테고리 조회
            category = self.get_keyword_category(keyword)
            if category:
                result['category'] = category
            
            result['success'] = True
            logger.info(f"키워드 분석 완료: {keyword} (볼륨: {monthly_volume}, 카테고리: {category})")
            
        except Exception as e:
            result['error_message'] = str(e)
            logger.error(f"키워드 분석 실패: {keyword}: {e}")
        
        return result
    
    
    def check_multiple_keywords_rank(self, keywords: List[str], product_id: str) -> List[RankingCheckDTO]:
        """여러 키워드의 순위를 한번에 검색"""
        results = []
        for keyword in keywords:
            try:
                result = self.check_product_rank(keyword, product_id)
                results.append(result)
            except Exception as e:
                logger.error(f"키워드 순위 검색 실패: {keyword}: {e}")
                # 실패한 경우도 결과에 포함
                failed_result = RankingCheckDTO(
                    success=False,
                    rank=RANK_OUT_OF_RANGE,
                    total_results=0,
                    error=str(e),
                    keyword=keyword,
                    product_id=product_id
                )
                results.append(failed_result)
        
        return results
    
    def analyze_keywords_for_tracking(self, keywords: List[str]) -> List[Dict[str, Any]]:
        """여러 키워드의 검색량과 카테고리를 한번에 분석"""
        results = []
        for keyword in keywords:
            try:
                result = self.analyze_keyword_for_tracking(keyword)
                results.append(result)
            except Exception as e:
                logger.error(f"키워드 분석 실패: {keyword}: {e}")
                # 실패한 경우도 결과에 포함
                failed_result = {
                    'keyword': keyword,
                    'monthly_volume': -1,
                    'category': '',
                    'success': False,
                    'error_message': str(e)
                }
                results.append(failed_result)
        
        return results
    
    def _clean_product_name(self, name: str) -> str:
        """상품명 정리 (HTML 태그 제거 등)"""
        return clean_product_name(name)
    
    def check_keyword_ranking(self, keyword: str, product_id: str) -> RankingCheckDTO:
        """키워드 순위 확인 (호호성 위한 alias - 향후 check_product_rank로 마이그레이션 후 제거)"""
        return self.check_product_rank(keyword, product_id)
    
    def analyze_keywords_batch(self, keywords: List[str]) -> Dict[str, Any]:
        """키워드 배치 월검색량 분석 (engine_local에서 이동)"""
        try:
            updated_count = 0
            failed_count = 0
            results = []
            
            for keyword in keywords:
                try:
                    analysis = self.analyze_keyword_for_tracking(keyword)
                    
                    if analysis['success']:
                        updated_count += 1
                        results.append({
                            'keyword': keyword,
                            'success': True,
                            'category': analysis['category'],
                            'monthly_volume': analysis['monthly_volume']
                        })
                    else:
                        failed_count += 1
                        results.append({
                            'keyword': keyword,
                            'success': False,
                            'error_message': analysis.get('error_message', '분석 실패')
                        })
                        
                except Exception as e:
                    failed_count += 1
                    results.append({
                        'keyword': keyword,
                        'success': False,
                        'error_message': str(e)
                    })
                    logger.error(f"키워드 '{keyword}' 처리 실패: {e}")
            
            return {
                'success': updated_count > 0,
                'updated_count': updated_count,
                'failed_count': failed_count,
                'total_count': len(keywords),
                'results': results
            }
            
        except Exception as e:
            logger.error(f"키워드 배치 분석 실패: {e}")
            return {
                'success': False,
                'updated_count': 0,
                'failed_count': len(keywords),
                'total_count': len(keywords),
                'results': [],
                'error_message': str(e)
            }
    
    def check_project_rankings_analysis(self, project, keywords: List) -> Dict[str, Any]:
        """프로젝트 전체 키워드 순위 확인 분석 (engine_local에서 이동)"""
        try:
            # 순위 확인 결과
            results = []
            success_count = 0
            
            for keyword_obj in keywords:
                result = self.check_keyword_ranking(keyword_obj.keyword, project.product_id)
                results.append(result)
                if result['success']:
                    success_count += 1
            
            return {
                'success': success_count > 0,
                'message': f"순위 확인 완료: {success_count}/{len(keywords)} 키워드",
                'results': results,
                'success_count': success_count,
                'total_count': len(keywords)
            }
            
        except Exception as e:
            logger.error(f"프로젝트 순위 확인 분석 실패: {e}")
            return {
                'success': False,
                'message': f"순위 확인 중 오류 발생: {e}",
                'results': [],
                'success_count': 0,
                'total_count': len(keywords) if keywords else 0
            }
    
    def process_single_keyword_ranking(self, keyword_obj, product_id: str) -> Tuple[Any, bool]:
        """단일 키워드 순위 확인 처리 (engine_local에서 이동)"""
        try:
            # 순위 확인
            result = self.check_keyword_ranking(keyword_obj.keyword, product_id)
            logger.info(f"순위 확인 결과: {keyword_obj.keyword} -> 순위: {result['rank']}, 성공: {result['success']}")
            return result, True
            
        except Exception as e:
            logger.error(f"키워드 {keyword_obj.keyword} 순위 확인 실패: {e}")
            failed_result = self._create_failed_ranking_result(keyword_obj.keyword, str(e))
            return failed_result, False
    
    def _create_failed_ranking_result(self, keyword: str, error: str) -> RankingCheckDTO:
        """실패한 순위 결과 생성"""
        return RankingCheckDTO(
            keyword=keyword,
            success=False,
            rank=RANK_OUT_OF_RANGE,
            error=error
        )
    
    def analyze_and_add_keyword(self, keyword: str) -> Dict[str, Any]:
        """키워드 분석 및 추가 로직 (engine_local에서 이동)"""
        try:
            # 키워드 분석 수행
            analysis = self.analyze_keyword_for_tracking(keyword)
            
            if analysis['success']:
                return {
                    'success': True,
                    'keyword': keyword,
                    'category': analysis['category'],
                    'monthly_volume': analysis['monthly_volume'],
                    'ready_for_db': True
                }
            else:
                return {
                    'success': False,
                    'keyword': keyword,
                    'category': '-',
                    'monthly_volume': -1,
                    'error_message': analysis.get('error_message', '분석 실패'),
                    'ready_for_db': False
                }
                
        except Exception as e:
            logger.error(f"키워드 '{keyword}' 분석/추가 로직 실패: {e}")
            return {
                'success': False,
                'keyword': keyword,
                'category': '-',
                'monthly_volume': -1,
                'error_message': str(e),
                'ready_for_db': False
            }
    
    def refresh_product_info_analysis(self, project) -> Dict[str, Any]:
        """프로젝트 상품 정보 새로고침 분석 (engine_local에서 이동)"""
        try:
            logger.info(f"프로젝트 정보 새로고침 시작: {project.current_name}")
            
            # 상품 정보 재조회
            product_info_dict = smart_product_search(project.current_name, project.product_id)
            
            if not product_info_dict:
                return {
                    'success': False,
                    'message': f'{project.current_name} 상품 정보를 찾을 수 없습니다.',
                    'changes': []
                }
            
            # 변경사항 분석은 engine에 위임
            from .engine_local import rank_tracking_engine
            from .models import ProductInfo
            
            # 변경사항 분석
            new_info = {
                'current_name': clean_product_name(product_info_dict.get('name', '')),
                'price': product_info_dict.get('price', 0),
                'category': product_info_dict.get('category', ''),
                'store_name': product_info_dict.get('store_name', ''),
            }
            
            # 변경사항 감지는 엔진에 위임 (순수 계산)
            changes_detected = rank_tracking_engine.detect_project_changes(project, new_info)
            
            # ProductInfo 객체 생성
            product_info = ProductInfo(
                product_id=product_info_dict.get('product_id', ''),
                name=new_info['current_name'],
                price=new_info['price'],
                category=new_info['category'],
                store_name=new_info['store_name'],
                description=product_info_dict.get('description', ''),
                image_url=product_info_dict.get('image_url', ''),
                url=product_info_dict.get('url', '')
            )
            
            return {
                'success': True,
                'new_info': new_info,
                'changes': changes_detected,
                'product_info': product_info
            }
            
        except Exception as e:
            logger.error(f"프로젝트 정보 새로고침 분석 실패: {e}")
            return {
                'success': False,
                'message': f'상품 정보 새로고침 중 오류가 발생했습니다: {str(e)}',
                'changes': []
            }
    
    def get_product_category_analysis(self, product_id: str) -> Dict[str, Any]:
        """상품 카테고리 조회 분석 (engine_local에서 이동)"""
        try:
            # smart_product_search를 통해 상품 정보 조회
            product_info = smart_product_search(f"상품ID_{product_id}", product_id)
            if product_info and 'category' in product_info:
                return {
                    'success': True,
                    'category': product_info['category'],
                    'product_id': product_id
                }
            
            return {
                'success': False,
                'category': '',
                'product_id': product_id,
                'error_message': '상품 정보 조회 실패'
            }
            
        except Exception as e:
            logger.error(f"상품 카테고리 분석 실패: {e}")
            return {
                'success': False,
                'category': '',
                'product_id': product_id,
                'error_message': str(e)
            }
    
    def process_keyword_info_analysis(self, keyword: str) -> Dict[str, Any]:
        """키워드 정보 분석 처리 (engine_local에서 이동)"""
        try:
            # 카테고리와 월검색량 조회
            category = self.get_keyword_category(keyword)
            monthly_volume = self.get_keyword_monthly_volume(keyword)
            
            return {
                'success': True,
                'category': category if category else '-',
                'monthly_volume': monthly_volume if monthly_volume is not None else -1,
                'keyword': keyword
            }
            
        except Exception as e:
            logger.error(f"키워드 정보 분석 실패: {keyword}: {e}")
            return {
                'success': False,
                'category': '-',
                'monthly_volume': -1,
                'keyword': keyword,
                'error_message': str(e)
            }


def to_export_row(keyword_row: dict) -> dict:
    """키워드 행 데이터를 엑셀 내보내기용 뷰모델로 변환"""
    return {
        "키워드": keyword_row.get("keyword", ""),
        "카테고리": keyword_row.get("category", "") or "-",
        "월검색량": format_monthly_volume(keyword_row.get("monthly_volume")),
        "현재순위": format_rank(keyword_row.get("rank_position")),
        "점검시각": format_datetime(keyword_row.get("search_date")),
    }


def to_display_row(keyword_row: dict) -> dict:
    """키워드 행 데이터를 UI 표시용 뷰모델로 변환"""
    return {
        "keyword": keyword_row.get("keyword", ""),
        "category": keyword_row.get("category", "") or "-",
        "monthly_volume_display": format_monthly_volume(keyword_row.get("monthly_volume")),
        "rank_display": format_rank(keyword_row.get("rank_position")),
        "search_date_display": format_datetime(keyword_row.get("search_date")),
        # 원본 데이터도 유지 (정렬/필터링용)
        "monthly_volume_raw": keyword_row.get("monthly_volume"),
        "rank_position_raw": keyword_row.get("rank_position"),
        "search_date_raw": keyword_row.get("search_date"),
    }


class RankTrackingExcelExporter:
    """순위 추적 Excel 내보내기 어댑터 (파일 I/O 담당)"""
    
    def get_default_filename(self, project_id: Optional[int] = None) -> str:
        """기본 파일명 생성"""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        
        if project_id:
            try:
                from .service import rank_tracking_service
                project = rank_tracking_service.get_project_by_id(project_id)
                if project:
                    # 파일명에 사용할 수 없는 문자 제거
                    safe_name = project.current_name.replace('/', '_').replace('\\', '_')
                    safe_name = safe_name.replace(':', '_').replace('*', '_')
                    safe_name = safe_name.replace('?', '_').replace('"', '_')
                    safe_name = safe_name.replace('<', '_').replace('>', '_')
                    safe_name = safe_name.replace('|', '_')
                    
                    return f"순위이력_{safe_name}_{timestamp}.xlsx"
            except Exception as e:
                logger.error(f"프로젝트명 조회 실패: {e}")
        
        return f"순위이력_데이터_{timestamp}.xlsx"
    
    def export_ranking_history_to_excel(self, project_id: int, file_path: str, ranking_data: list = None) -> bool:
        """순위 이력을 Excel로 내보내기 (원본과 동일)"""
        try:
            from .service import rank_tracking_service
            
            # 프로젝트 정보 조회
            project = rank_tracking_service.get_project_by_id(project_id)
            if not project:
                logger.error(f"프로젝트를 찾을 수 없습니다: {project_id}")
                return False
            
            # 키워드 정보 조회
            keywords = rank_tracking_service.get_project_keywords(project_id)
            if not keywords:
                logger.error("키워드가 없습니다")
                return False
            
            # 프로젝트 순위 개요 데이터 사용
            overview = rank_tracking_service.get_project_overview(project_id)
            all_dates = overview.get('dates', [])[:10]  # 최대 10개 날짜
            keywords_data = overview.get('keywords', {})
            
            logger.info(f"디버깅: 키워드 수 = {len(keywords)}")
            logger.info(f"디버깅: 전체 날짜 수 = {len(all_dates)}, 날짜들 = {all_dates}")
            
            # 키워드별 순위 데이터 구성
            keyword_ranking_data = []
            for keyword_obj in keywords:
                keyword_data = keywords_data.get(keyword_obj.keyword, {})
                rankings = keyword_data.get('rankings', {})
                
                logger.info(f"디버깅: 키워드 '{keyword_obj.keyword}' 순위 이력 수 = {len(rankings)}")
                
                # 날짜별 순위 매핑 (overview 데이터 형식에 맞춤)
                rank_by_date = {}
                for date in all_dates:
                    if date in rankings:
                        rank = rankings[date].get('rank', 999)
                        rank_by_date[date] = rank
                        logger.info(f"디버깅: 순위 데이터 - 키워드: {keyword_obj.keyword}, 날짜: {date}, 순위: {rank}")
                
                keyword_ranking_data.append({
                    'keyword': keyword_obj.keyword,
                    'category': keyword_obj.category or '-',
                    'monthly_volume': keyword_obj.monthly_volume if keyword_obj.monthly_volume is not None else -1,
                    'rank_by_date': rank_by_date
                })
            
            # 날짜 정렬 (최신순) 및 형식 변환
            sorted_dates = []
            formatted_dates = []
            for date in all_dates:
                # 날짜를 MM/DD HH:MM 형식으로 변환
                try:
                    if isinstance(date, str):
                        dt = _to_dt(date)
                        if dt:
                            formatted_date = dt.strftime("%m/%d %H:%M")
                            sorted_dates.append(date)  # 원본 날짜 (키 매칭용)
                            formatted_dates.append(formatted_date)  # 표시용 날짜
                            logger.info(f"디버깅: 날짜 변환 - {date} -> {formatted_date}")
                        else:
                            logger.warning(f"디버깅: 날짜 파싱 실패 - {date}")
                            continue
                except Exception as e:
                    logger.warning(f"디버깅: 날짜 변환 실패 - {date}: {e}")
                    continue
            
            logger.info(f"디버깅: 최종 날짜 수 = {len(sorted_dates)}, 변환된 날짜들 = {formatted_dates}")
            
            # 엑셀 데이터 구성
            excel_data = []
            
            # 1. 기본정보 섹션 (사진과 똑같이)
            excel_data.extend([
                [f"📊 {project.current_name}", "", "", "", "", "", "", "", ""],
                ["", "", "", "", "", "", "", "", ""],
                ["상품 ID", project.product_id, "", "", "", "", "", "", ""],
                ["상품명", project.current_name, "", "", "", "", "", "", ""],
                ["스토어명", project.store_name or "-", "", "", "", "", "", "", ""],
                ["가격", f"{project.price:,}원" if project.price else "-", "", "", "", "", "", "", ""],
                ["카테고리", project.category or "-", "", "", "", "", "", "", ""],
                ["등록일", self._format_date(project.created_at) if project.created_at else "-", "", "", "", "", "", "", ""],
                ["", "", "", "", "", "", "", "", ""],
                ["", "", "", "", "", "", "", "", ""],
                ["🔍 키워드 순위 현황", "", "", "", "", "", "", "", ""]
            ])
            
            # 2. 키워드 순위 테이블 헤더 (표시용 날짜 사용)
            header_row = ["키워드", "카테고리", "월검색량"]
            header_row.extend(formatted_dates)  # 변환된 날짜들 추가
            excel_data.append(header_row)
            
            # 3. 키워드별 순위 데이터
            for kw_data in keyword_ranking_data:
                # 새로운 포맷터 사용
                volume_display = format_monthly_volume(kw_data['monthly_volume'])
                
                data_row = [
                    kw_data['keyword'],
                    kw_data['category'],
                    volume_display
                ]
                
                # 각 날짜별 순위 추가 (원본 날짜로 키 매칭)
                for date in sorted_dates:
                    rank = kw_data['rank_by_date'].get(date, "")
                    if rank:
                        if rank == 999 or rank > 200:  # RANK_OUT_OF_RANGE 또는 200위 초과
                            data_row.append("200+")
                        else:
                            data_row.append(f"{rank}위")
                    else:
                        data_row.append("")
                
                excel_data.append(data_row)
            
            # 엑셀 파일 생성
            success = self._create_excel_file(file_path, excel_data)
            
            if success:
                logger.info(f"순위 이력 엑셀 파일 생성 완료: {file_path}")
            else:
                logger.error("순위 이력 엑셀 파일 생성 실패")
            
            return success
            
        except Exception as e:
            logger.error(f"순위 이력 엑셀 저장 중 오류: {e}")
            return False
    
    def export_multiple_projects_to_excel(self, project_ids: List[int], file_path: str) -> bool:
        """여러 프로젝트를 엑셀로 저장 (단일 저장을 시트별로 분할)"""
        try:
            import openpyxl
            
            # 워크북 생성
            workbook = openpyxl.Workbook()
            
            # 기본 시트 제거
            if 'Sheet' in workbook.sheetnames:
                del workbook['Sheet']
            
            # 각 프로젝트별로 시트 생성
            for i, project_id in enumerate(project_ids, 1):
                try:
                    logger.info(f"다중 프로젝트 내보내기: 프로젝트 {project_id} 처리 시작")
                    
                    # 임시 파일에 단일 프로젝트 저장 (temp 폴더 사용)
                    import tempfile
                    import os
                    temp_dir = tempfile.gettempdir()
                    temp_file = os.path.join(temp_dir, f"temp_project_{project_id}.xlsx")
                    success = self.export_ranking_history_to_excel(project_id, temp_file)
                    
                    if success:
                        logger.info(f"프로젝트 {project_id} 임시 파일 생성 성공: {temp_file}")
                        
                        # 임시 파일을 워크북에 시트로 추가
                        temp_workbook = openpyxl.load_workbook(temp_file)
                        source_sheet = temp_workbook.active
                        
                        # 새 시트 생성
                        target_sheet = workbook.create_sheet(title=f"Sheet{i}")
                        
                        # 데이터와 스타일 안전한 복사
                        for row in source_sheet.iter_rows():
                            for cell in row:
                                new_cell = target_sheet.cell(row=cell.row, column=cell.column, value=cell.value)
                                
                                # 스타일 안전한 복사 (StyleProxy 오류 방지)
                                try:
                                    if cell.font:
                                        from openpyxl.styles import Font
                                        new_cell.font = Font(
                                            name=cell.font.name,
                                            size=cell.font.size,
                                            bold=cell.font.bold,
                                            italic=cell.font.italic,
                                            color=cell.font.color
                                        )
                                except:
                                    pass
                                    
                                try:
                                    if cell.fill:
                                        from openpyxl.styles import PatternFill
                                        new_cell.fill = PatternFill(
                                            start_color=cell.fill.start_color,
                                            end_color=cell.fill.end_color,
                                            fill_type=cell.fill.fill_type
                                        )
                                except:
                                    pass
                                    
                                try:
                                    if cell.alignment:
                                        from openpyxl.styles import Alignment
                                        new_cell.alignment = Alignment(
                                            horizontal=cell.alignment.horizontal,
                                            vertical=cell.alignment.vertical
                                        )
                                except:
                                    pass
                                    
                                try:
                                    if cell.number_format:
                                        new_cell.number_format = cell.number_format
                                except:
                                    pass
                        
                        # 컬럼 너비 복사
                        for col_letter, dimension in source_sheet.column_dimensions.items():
                            target_sheet.column_dimensions[col_letter].width = dimension.width
                        
                        # 임시 워크북 완전히 닫기
                        temp_workbook.close()
                        
                        # 파일 잠금 해제를 위한 대기
                        import time
                        time.sleep(0.3)
                        
                        # 임시 파일 안전하게 삭제 (조용히)
                        import os
                        try:
                            if os.path.exists(temp_file):
                                os.remove(temp_file)
                                logger.debug(f"임시 파일 삭제 성공: {temp_file}")
                        except:
                            # 삭제 실패는 로그 출력하지 않음 (시스템이 알아서 정리함)
                            pass
                        
                        logger.info(f"프로젝트 {project_id} 시트 생성 완료")
                    else:
                        logger.error(f"프로젝트 {project_id} 단일 저장 실패")
                        
                except Exception as e:
                    logger.error(f"프로젝트 {project_id} 처리 중 오류: {e}")
                    continue
            
            if len(workbook.sheetnames) == 0:
                logger.error("저장할 프로젝트 데이터가 없습니다")
                workbook.close()
                return False
            
            # 파일 저장
            workbook.save(file_path)
            workbook.close()
            
            logger.info(f"다중 프로젝트 엑셀 파일 생성 완료: {file_path}")
            return True
            
        except Exception as e:
            logger.error(f"다중 프로젝트 엑셀 저장 중 오류: {e}")
            return False
    
    def _format_date(self, date_value):
        """날짜 형식을 안전하게 변환"""
        try:
            if isinstance(date_value, str):
                # 문자열인 경우 datetime으로 변환
                dt = _to_dt(date_value)
                return dt.strftime("%Y-%m-%d") if dt else str(date_value)
            elif hasattr(date_value, 'strftime'):
                # datetime 객체인 경우
                return date_value.strftime("%Y-%m-%d")
            else:
                # 기타 경우는 문자열로 변환
                return str(date_value)
        except Exception as e:
            logger.warning(f"날짜 형식 변환 실패: {e}")
            return str(date_value) if date_value else "-"
    
    def _create_excel_file(self, file_path: str, excel_data: list) -> bool:
        """엑셀 파일 생성"""
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            
            # 워크북 생성
            workbook = openpyxl.Workbook()
            worksheet = workbook.active
            worksheet.title = "순위이력"
            
            # 데이터 입력 및 정렬 가능하도록 처리
            for row_idx, row_data in enumerate(excel_data, 1):
                for col_idx, cell_value in enumerate(row_data, 1):
                    # 월검색량과 순위 컬럼은 숫자로 저장하여 정렬 가능하게 함
                    if row_idx > 12 and col_idx == 3:  # 월검색량 컬럼
                        # 새로운 포맷터 기반 처리
                        try:
                            if isinstance(cell_value, str):
                                if cell_value == "미수집" or cell_value == "N/A":
                                    # 미수집/N/A는 문자열로 저장
                                    cell = worksheet.cell(row=row_idx, column=col_idx, value=cell_value)
                                elif cell_value == "0":
                                    # 검색량 0일 때는 숫자 0으로 저장
                                    cell = worksheet.cell(row=row_idx, column=col_idx, value=0)
                                    cell.number_format = '#,##0'
                                elif cell_value.replace(',', '').isdigit():
                                    numeric_value = int(cell_value.replace(',', ''))
                                    cell = worksheet.cell(row=row_idx, column=col_idx, value=numeric_value)
                                    cell.number_format = '#,##0'
                                else:
                                    cell = worksheet.cell(row=row_idx, column=col_idx, value=cell_value)
                            elif isinstance(cell_value, int):
                                cell = worksheet.cell(row=row_idx, column=col_idx, value=cell_value)
                                cell.number_format = '#,##0'
                            else:
                                cell = worksheet.cell(row=row_idx, column=col_idx, value=cell_value)
                        except:
                            cell = worksheet.cell(row=row_idx, column=col_idx, value=cell_value)
                    elif row_idx > 12 and col_idx > 3:  # 순위 컬럼들
                        # 순위를 숫자로 변환하여 정렬 가능하게 함
                        try:
                            if isinstance(cell_value, str):
                                if "200+" in cell_value:
                                    cell = worksheet.cell(row=row_idx, column=col_idx, value=201)  # 정렬용
                                    # 200+ 표시를 위한 number format 설정
                                    cell.number_format = '"200+"'
                                elif "위" in cell_value:
                                    rank_num = int(cell_value.replace("위", ""))
                                    cell = worksheet.cell(row=row_idx, column=col_idx, value=rank_num)
                                    # N위 표시를 위한 number format 설정
                                    cell.number_format = '0"위"'
                                else:
                                    cell = worksheet.cell(row=row_idx, column=col_idx, value=cell_value)
                            else:
                                cell = worksheet.cell(row=row_idx, column=col_idx, value=cell_value)
                        except:
                            cell = worksheet.cell(row=row_idx, column=col_idx, value=cell_value)
                    else:
                        cell = worksheet.cell(row=row_idx, column=col_idx, value=cell_value)
                    
                    # 스타일 적용
                    if row_idx == 1:  # 제목 행
                        cell.font = Font(size=14, bold=True)
                        cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
                        cell.font = Font(color="FFFFFF", size=14, bold=True)
                    elif row_idx == 11:  # 키워드 순위 현황 헤더 (11번째 행)
                        cell.font = Font(size=12, bold=True)
                        cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
                    elif row_idx == 12:  # 테이블 헤더 (12번째 행이 실제 헤더)
                        cell.font = Font(color="FFFFFF", bold=True)
                        cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                    elif row_idx > 12:  # 데이터 행 (12번째 행 이후)
                        if col_idx <= 3:  # 키워드, 카테고리, 월검색량 컬럼
                            cell.alignment = Alignment(horizontal="left", vertical="center")
                        else:  # 순위 컬럼들
                            cell.alignment = Alignment(horizontal="center", vertical="center")
                            # 순위에 따른 색상 적용 (숫자 값 기준)
                            if isinstance(cell.value, (int, float)):
                                rank_num = int(cell.value)
                                if rank_num <= 10:
                                    cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # 연한 초록색
                                elif rank_num <= 50:
                                    cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")  # 연한 노란색
                                elif rank_num <= 200:
                                    cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # 연한 빨간색
                                else:  # 200+
                                    cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # 연한 빨간색
            
            # 컬럼 너비 설정
            for col_idx in range(1, worksheet.max_column + 1):
                column_letter = worksheet.cell(row=1, column=col_idx).column_letter
                
                if col_idx == 1:  # 키워드 컬럼
                    worksheet.column_dimensions[column_letter].width = 20
                elif col_idx == 2:  # 카테고리 컬럼
                    worksheet.column_dimensions[column_letter].width = 30
                elif col_idx == 3:  # 월검색량 컬럼
                    worksheet.column_dimensions[column_letter].width = 12
                else:  # 순위 컬럼들 (기존 대비 1.5배)
                    worksheet.column_dimensions[column_letter].width = 15  # 기본 10 → 15로 1.5배
            
            # 파일 저장
            workbook.save(file_path)
            return True
            
        except Exception as e:
            logger.error(f"엑셀 파일 생성 중 오류: {e}")
            return False


# 전역 어댑터 인스턴스
rank_tracking_adapter = RankTrackingAdapter()
rank_tracking_excel_exporter = RankTrackingExcelExporter()